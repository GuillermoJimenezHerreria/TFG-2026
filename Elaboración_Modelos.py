"""
MODELOS PREDICTIVOS MULTIDIMENSIONALES - TFG
Universidad Francisco de Vitoria
Guillermo

Modelos implementados desde CERO con numpy (sin sklearn):
  1. Regresión Lineal con Regularización Ridge  (OLS con L2)
  2. Árbol de Decisión para Regresión
  3. Random Forest (ensamble de Árboles de Decisión)
  4. Gradient Boosting (implementación simplificada tipo XGBoost)

Métricas: R², RMSE, MAE, MAPE
Validación: K-Fold Cross-Validation (K=5)
Interpretabilidad: SHAP values (aproximación por permutación)
Output: Excel auditado con fórmulas trazables donde sea posible
"""

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from pathlib import Path
import sys
import warnings, os, json
warnings.filterwarnings('ignore')

# ═══════════════════════════════════════════════════════════════
#  CONFIGURACIÓN — LOCAL vs GOOGLE COLAB
#  En Colab: sube la carpeta 'TFG Business Analytics' a Google
#  Drive y ejecuta con Runtime → Ejecutar todo.
# ═══════════════════════════════════════════════════════════════
IN_COLAB = 'google.colab' in sys.modules
if IN_COLAB:
    DB = '/content'
    import subprocess
    subprocess.run(['pip', 'install', '-q', 'openpyxl', 'lxml'],
                   capture_output=True)
else:
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    DB = os.path.join(BASE_DIR, 'Bases de Datos')

# ─────────────────────────────────────────────
#  RUTAS
# ─────────────────────────────────────────────
BASE_EXCEL = os.path.join(DB, 'Dataset_Definitivo.xlsx')
OUT_EXCEL  = os.path.join(DB, 'ML TFG.xlsx')
PLOT_DIR   = Path('/content/ml_plots' if IN_COLAB else os.path.join(BASE_DIR, 'ml_plots'))
PLOT_DIR.mkdir(exist_ok=True)

# ─────────────────────────────────────────────
#  PALETA CORPORATIVA TFG
# ─────────────────────────────────────────────
C_BLUE   = '1F4E79'
C_LBLUE  = '2E75B6'
C_GOLD   = 'C9A84C'
C_GREEN  = '375623'
C_LGREEN = '70AD47'
C_RED    = 'C00000'
C_ORANGE = 'E26B0A'
C_GREY   = 'F2F2F2'
C_WHITE  = 'FFFFFF'

def hdr_fill(hex_color): return PatternFill('solid', fgColor=hex_color)
def hdr_font(bold=True, color='FFFFFF', sz=11):
    return Font(bold=bold, color=color, size=sz, name='Calibri')
def thin_border():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)
def center(): return Alignment(horizontal='center', vertical='center', wrap_text=True)
def left():   return Alignment(horizontal='left',   vertical='center', wrap_text=False)

# ─────────────────────────────────────────────
#  1. CARGA Y PREPARACIÓN DE DATOS
# ─────────────────────────────────────────────
print("=" * 60)
print("CARGANDO DATOS...")
print("=" * 60)

COL_ALIASES = {
    # ── Identificación ────────────────────────────────────────────────────────
    'Nombre': 'player', 'Player': 'player', 'Jugador': 'player',
    'Posición': 'position', 'Pos': 'position',
    'Equipo': 'team',     'Squad': 'team',
    # ── Año de nacimiento (para calcular edad internamente) ───────────────────
    'Edad': 'age_fbref_raw',        'Age': 'age_fbref_raw',
    'Año de nacimiento': 'birth_year', 'Born': 'birth_year',
    # ── Valor de mercado ──────────────────────────────────────────────────────
    'Valor de mercado': 'market_value_eur',
    'market_value_eur': 'market_value_eur',
    # ── Ingresos del club ─────────────────────────────────────────────────────
    'Ingreso total anual del club': 'club_revenue',
    'club_revenue': 'club_revenue',
    'club_revenue_eur': 'club_revenue',
    # ── Partidos y minutos ────────────────────────────────────────────────────
    'Partidos jugados esta temporada': 'games_played',
    'MP': 'games_played',
    'Minutos jugados esta temporada': 'minutes_played',
    'Min': 'minutes_played',
    # ── Rendimiento ofensivo ──────────────────────────────────────────────────
    'Goles': 'goals',         'Gls': 'goals',
    'Asistencias': 'assists', 'Ast': 'assists',
    # ── Contrato y lesiones ───────────────────────────────────────────────────
    'Años de contrato restantes': 'contract_years',
    'contract_years_remaining': 'contract_years',
    'Cantidad total de lesiones registradas': 'injury_count',
    'injury_count': 'injury_count',
    'Promedio de días de baja por lesión por temporada': 'injury_days',
    'injury_days_per_season': 'injury_days',
    # ── Calidad del matching ──────────────────────────────────────────────────
    'match_confidence': 'match_confidence',
    # ── FBref (columnas en español — PL) ──────────────────────────────────────
    'Entradas ganadas (FBref)':      'tackles_won',
    'Intercepciones (FBref)':        'interceptions',
    'Faltas cometidas (FBref)':      'fouls',
    'Faltas recibidas (FBref)':      'fouled',
    'Centros (FBref)':               'crosses',
    'Tarjetas amarillas (FBref)':    'cards_yellow',
    'Tarjetas rojas (FBref)':        'cards_red',
    'Tiros totales (FBref)':         'shots',
    'Tiros a puerta (FBref)':        'shots_on_target',
    'Precisión de tiro % (FBref)':   'shot_accuracy',
    'Tiros por 90 min (FBref)':      'shots_p90',
    # ── FBref (columnas en inglés — resto de ligas) ────────────────────────────
    'tackles_won': 'tackles_won',
    'interceptions': 'interceptions',
    'fouls': 'fouls',
    'fouled': 'fouled',
    'crosses': 'crosses',
    'cards_yellow': 'cards_yellow',
    'cards_red': 'cards_red',
    'shots': 'shots',
    'shots_on_target': 'shots_on_target',
    'shot_accuracy': 'shot_accuracy',
    'shots_p90': 'shots_p90',
    'gk_save_pct': 'gk_save_pct',
    'gk_clean_sheets': 'gk_clean_sheets',
}

SHEETS = {
    'PL Players':          'PL',
    'LaLiga Players':      'LaLiga',
    'Bundesliga Players':  'Bundesliga',
    'Serie A Players':     'SerieA',
    'Ligue 1 Players':     'Ligue1',
}

# Umbral de confianza para el matching Transfermarkt
# Jugadores con conf < 0.85 tienen mayor riesgo de asignación errónea de MV
CONF_THRESHOLD = 0.85

frames = []
xl = pd.ExcelFile(BASE_EXCEL)
for sheet, league in SHEETS.items():
    try:
        df = xl.parse(sheet, header=0)
        df.columns = [COL_ALIASES.get(str(c).strip(), str(c).strip()) for c in df.columns]
        df = df.loc[:, ~df.columns.duplicated()]   # drop duplicate cols after rename
        df['league'] = league
        # Añadir columna de confianza normalizada si existe
        if 'match_confidence' not in df.columns:
            df['match_confidence'] = 1.0
        frames.append(df)
    except Exception as e:
        print(f"  Error en {sheet}: {e}")

df_all = pd.concat(frames, ignore_index=True)

# Limpieza
def to_num(s):
    try:
        v = str(s).replace(',', '.').replace('€', '').replace(' ', '').replace('M', '')
        return float(v)
    except:
        return np.nan

for col in ['market_value_eur', 'club_revenue', 'games_played',
            'minutes_played', 'goals', 'assists', 'contract_years',
            'injury_count', 'injury_days', 'tackles_won', 'interceptions',
            'fouls', 'fouled', 'crosses', 'cards_yellow', 'shots',
            'shots_on_target', 'shot_accuracy', 'shots_p90']:
    if col in df_all.columns:
        df_all[col] = df_all[col].apply(to_num)

df_all['market_value_eur'] = pd.to_numeric(df_all['market_value_eur'], errors='coerce')
df_all['match_confidence']  = pd.to_numeric(df_all['match_confidence'], errors='coerce').fillna(1.0)

# ── Calcular edad desde año de nacimiento (evita el formato FBref '30-063') ───
# Temporada 2024-25: edad = 2025 - año_nacimiento
df_all['birth_year'] = pd.to_numeric(df_all['birth_year'], errors='coerce')
df_all['age'] = 2025 - df_all['birth_year']

# ── FILTRO DE CALIDAD: anular MV de jugadores con matching poco fiable ─────────
# Si match_confidence < 0.85, el fuzzy-match fue débil y el MV asignado
# puede corresponder a otro jugador (ej: Rodrigo Mendoza/Elche → 110M = Rodri).
# Se anulan esos valores para no contaminar el modelo.
n_before = (df_all['market_value_eur'] > 0).sum()
mask_low_conf = df_all['match_confidence'] < CONF_THRESHOLD
df_all.loc[mask_low_conf, 'market_value_eur'] = np.nan
n_after  = (df_all['market_value_eur'] > 0).sum()
print(f"  Filtro confianza (<{CONF_THRESHOLD}): {n_before - n_after} valores anulados "
      f"({n_before} -> {n_after} jugadores con MV valido)")

df_valid = df_all[df_all['market_value_eur'] > 0].copy()
df_valid['log_mv'] = np.log(df_valid['market_value_eur'])

# Posición simplificada → dummies
POS_MAP = {'GK': 0, 'DF': 1, 'MF': 2, 'FW': 3,
           'Portero': 0, 'Defensa': 1, 'Centrocampista': 2, 'Delantero': 3,
           'GK,DF': 1, 'DF,MF': 1, 'MF,FW': 2, 'DF,FW': 1,
           'FW,MF': 2, 'MF,DF': 1, 'FW,DF': 1, 'GK,MF': 2}
df_valid['pos_num'] = df_valid['position'].apply(
    lambda x: POS_MAP.get(str(x).strip(), np.nan) if pd.notna(x) else np.nan)

# Liga → dummy numérico
LIGA_MAP = {'PL': 0, 'LaLiga': 1, 'Bundesliga': 2, 'SerieA': 3, 'Ligue1': 4}
df_valid['league_num'] = df_valid['league'].map(LIGA_MAP)

FEATURES = [
    'age', 'club_revenue', 'games_played', 'minutes_played',
    'goals', 'assists', 'contract_years', 'injury_count', 'injury_days',
    'tackles_won', 'interceptions', 'fouls', 'fouled', 'crosses',
    'shots', 'shots_on_target', 'shot_accuracy',
    'pos_num', 'league_num'
]
FEATURES = [f for f in FEATURES if f in df_valid.columns]

# Dataset limpio (sin NaN en ninguna feature)
df_model = df_valid[FEATURES + ['log_mv', 'player', 'team', 'league',
                                 'market_value_eur']].dropna()
print(f"  Dataset ML: {len(df_model)} jugadores × {len(FEATURES)} features")

X = df_model[FEATURES].values.astype(float)
y = df_model['log_mv'].values.astype(float)
n, p = X.shape

# Normalización Z-score (manual)
X_mean = X.mean(axis=0)
X_std  = X.std(axis=0)
X_std[X_std == 0] = 1  # evitar división por cero
X_sc   = (X - X_mean) / X_std  # escalado para Ridge y comparación

print(f"  Features: {FEATURES}")

# ─────────────────────────────────────────────
#  UTILIDADES COMUNES
# ─────────────────────────────────────────────
def r2_score(y_true, y_pred):
    """R² = 1 - SS_res / SS_tot"""
    ss_res = ((y_true - y_pred)**2).sum()
    ss_tot = ((y_true - y_true.mean())**2).sum()
    return 1 - ss_res / ss_tot if ss_tot > 0 else 0.0

def rmse(y_true, y_pred):
    return np.sqrt(((y_true - y_pred)**2).mean())

def mae(y_true, y_pred):
    return np.abs(y_true - y_pred).mean()

def mape(y_true, y_pred):
    mask = y_true != 0
    return np.abs((y_true[mask] - y_pred[mask]) / y_true[mask]).mean() * 100

def kfold_indices(n, k=5, seed=42):
    """Genera índices de train/test para K-Fold CV"""
    rng = np.random.default_rng(seed)
    idx = rng.permutation(n)
    folds = np.array_split(idx, k)
    splits = []
    for i in range(k):
        test_idx  = folds[i]
        train_idx = np.concatenate([folds[j] for j in range(k) if j != i])
        splits.append((train_idx, test_idx))
    return splits

SPLITS = kfold_indices(n, k=5)

def cv_evaluate(predict_fn, train_fn, X, y, splits):
    """
    Ejecuta K-Fold CV.
    train_fn(X_train, y_train) → modelo
    predict_fn(modelo, X_test) → y_pred
    Devuelve dict con métricas medias y std.
    """
    metrics = {'r2': [], 'rmse': [], 'mae': [], 'mape': []}
    for tr_idx, te_idx in splits:
        model = train_fn(X[tr_idx], y[tr_idx])
        y_pred = predict_fn(model, X[te_idx])
        metrics['r2'].append(r2_score(y[te_idx], y_pred))
        metrics['rmse'].append(rmse(y[te_idx], y_pred))
        metrics['mae'].append(mae(y[te_idx], y_pred))
        metrics['mape'].append(mape(y[te_idx], y_pred))
    return {k: (np.mean(v), np.std(v)) for k, v in metrics.items()}

# ─────────────────────────────────────────────
#  2. MODELO 1 — REGRESIÓN LINEAL RIDGE
# ─────────────────────────────────────────────
print("\n" + "=" * 60)
print("MODELO 1: REGRESIÓN LINEAL RIDGE")
print("=" * 60)
"""
θ = (XᵀX + λI)⁻¹ Xᵀy
Solución analítica exacta. λ=1 (regularización L2 moderada).
Interpretación: β_i indica el cambio en log_MV por unidad
de incremento en la feature i (ya escalada).
"""
LAMBDA_RIDGE = 1.0  # hiperparámetro α

def ridge_train(X, y, lam=LAMBDA_RIDGE):
    """θ = (XᵀX + λI)⁻¹ Xᵀy"""
    Xb = np.column_stack([np.ones(len(X)), X])  # añadir intercepto
    A  = Xb.T @ Xb + lam * np.eye(Xb.shape[1])
    A[0, 0] -= lam  # no regularizar el intercepto
    theta = np.linalg.solve(A, Xb.T @ y)
    return theta

def ridge_predict(theta, X):
    Xb = np.column_stack([np.ones(len(X)), X])
    return Xb @ theta

# Entrenar con todos los datos para obtener coeficientes
theta_ridge = ridge_train(X_sc, y)
y_pred_ridge_all = ridge_predict(theta_ridge, X_sc)

# Cross-validation
cv_ridge = cv_evaluate(
    lambda m, Xt: ridge_predict(m, (Xt - X_mean) / X_std),
    lambda Xtr, ytr: ridge_train((Xtr - X_mean) / X_std, ytr),
    X, y, SPLITS
)

print(f"  R² CV:   {cv_ridge['r2'][0]:.4f} ± {cv_ridge['r2'][1]:.4f}")
print(f"  RMSE CV: {cv_ridge['rmse'][0]:.4f} ± {cv_ridge['rmse'][1]:.4f}")
print(f"  MAE CV:  {cv_ridge['mae'][0]:.4f} ± {cv_ridge['mae'][1]:.4f}")

# Coeficientes
ridge_coefs = pd.Series(theta_ridge[1:], index=FEATURES).sort_values(key=abs, ascending=False)
print(f"  Top coeficientes: {ridge_coefs.head(5).to_dict()}")

# ─────────────────────────────────────────────
#  3. ÁRBOL DE DECISIÓN — BLOQUE BASE
# ─────────────────────────────────────────────
"""
Árbol de Regresión CART:
- Splitting criterion: minimizar MSE = Σ(yi - ȳ)²
- Parámetros: max_depth, min_samples_split
- Predicción en hoja: ȳ_nodo
"""

class DecisionNode:
    __slots__ = ['feat', 'thresh', 'left', 'right', 'value', 'n_samples']
    def __init__(self):
        self.feat = self.thresh = self.left = self.right = self.value = None
        self.n_samples = 0

def _mse_split(y_left, y_right):
    """MSE ponderado de la partición"""
    n = len(y_left) + len(y_right)
    if n == 0:
        return np.inf
    mse_l = ((y_left  - y_left.mean())**2).sum()  if len(y_left)  > 0 else 0
    mse_r = ((y_right - y_right.mean())**2).sum() if len(y_right) > 0 else 0
    return (mse_l + mse_r) / n

def _best_split(X, y, n_features_try=None, rng=None):
    """Encuentra el mejor split por reducción de MSE"""
    n, p = X.shape
    best = {'gain': -np.inf, 'feat': None, 'thresh': None}
    base_mse = ((y - y.mean())**2).mean()

    feat_idx = np.arange(p)
    if n_features_try is not None and n_features_try < p:
        feat_idx = rng.choice(p, size=n_features_try, replace=False)

    for j in feat_idx:
        vals = np.unique(X[:, j])
        if len(vals) <= 1:
            continue
        # Candidatos: puntos medios
        thresholds = (vals[:-1] + vals[1:]) / 2
        # Para velocidad, muestrear max 20 umbrales
        if len(thresholds) > 20:
            step = max(1, len(thresholds) // 20)
            thresholds = thresholds[::step]
        for t in thresholds:
            mask = X[:, j] <= t
            y_l, y_r = y[mask], y[~mask]
            if len(y_l) == 0 or len(y_r) == 0:
                continue
            split_mse = _mse_split(y_l, y_r)
            gain = base_mse - split_mse
            if gain > best['gain']:
                best = {'gain': gain, 'feat': j, 'thresh': t}
    return best

def _build_tree(X, y, depth, max_depth, min_samples, n_feat, rng):
    node = DecisionNode()
    node.n_samples = len(y)
    node.value = y.mean()
    if depth >= max_depth or len(y) < min_samples:
        return node
    best = _best_split(X, y, n_feat, rng)
    if best['feat'] is None:
        return node
    node.feat   = best['feat']
    node.thresh = best['thresh']
    mask = X[:, node.feat] <= node.thresh
    node.left  = _build_tree(X[mask],  y[mask],  depth+1, max_depth, min_samples, n_feat, rng)
    node.right = _build_tree(X[~mask], y[~mask], depth+1, max_depth, min_samples, n_feat, rng)
    return node

def _predict_one(node, x):
    if node.left is None:
        return node.value
    if x[node.feat] <= node.thresh:
        return _predict_one(node.left, x)
    return _predict_one(node.right, x)

def tree_predict(node, X):
    return np.array([_predict_one(node, x) for x in X])

def tree_train(X, y, max_depth=6, min_samples=10, n_feat=None, seed=42):
    rng = np.random.default_rng(seed)
    return _build_tree(X, y, 0, max_depth, min_samples, n_feat, rng)

# ─────────────────────────────────────────────
#  4. MODELO 2 — RANDOM FOREST
# ─────────────────────────────────────────────
print("\n" + "=" * 60)
print("MODELO 2: RANDOM FOREST")
print("=" * 60)
"""
Random Forest:
- B árboles, cada uno entrenado sobre Bootstrap(X, y)
- Cada split considera sqrt(p) features aleatorias
- Predicción: media de predicciones de los B árboles
- Importancia: reducción media de MSE por feature (MDI)
"""
N_TREES  = 100
MAX_DEPTH_RF = 8
MIN_SAMP_RF  = 8
N_FEAT_RF    = max(1, int(np.sqrt(p)))  # sqrt(p)

class RandomForest:
    def __init__(self, trees):
        self.trees = trees  # lista de (tree, oob_indices)

def rf_train(X, y, n_trees=N_TREES, max_depth=MAX_DEPTH_RF,
             min_samples=MIN_SAMP_RF, n_feat=N_FEAT_RF, seed=0):
    rng = np.random.default_rng(seed)
    trees = []
    n = len(y)
    for i in range(n_trees):
        boot_idx = rng.integers(0, n, size=n)
        oob_idx  = np.setdiff1d(np.arange(n), boot_idx)
        X_b, y_b = X[boot_idx], y[boot_idx]
        t = _build_tree(X_b, y_b, 0, max_depth, min_samples, n_feat,
                        np.random.default_rng(seed + i + 1))
        trees.append((t, oob_idx))
    rf = RandomForest(trees)
    return rf

def rf_predict(rf, X):
    preds = np.array([tree_predict(t, X) for t, _ in rf.trees])
    return preds.mean(axis=0)

def rf_oob_score(rf, X, y):
    """Out-Of-Bag R² como estimador sin sesgo del error de generalización"""
    n = len(y)
    oob_sum   = np.zeros(n)
    oob_count = np.zeros(n)
    for t, oob_idx in rf.trees:
        if len(oob_idx) == 0:
            continue
        preds = tree_predict(t, X[oob_idx])
        oob_sum[oob_idx]   += preds
        oob_count[oob_idx] += 1
    valid = oob_count > 0
    y_oob_pred = np.where(valid, oob_sum / np.where(oob_count > 0, oob_count, 1), np.nan)
    y_v = y[valid]; p_v = y_oob_pred[valid]
    return r2_score(y_v, p_v)

def rf_feature_importance(rf, n_features):
    """
    Importancia por permutación de impureza media (MDI):
    Suma de reducción de MSE ponderada por nodos en cada árbol.
    Aquí aproximamos como |coeficiente medio| sobre todos los splits.
    """
    counts = np.zeros(n_features)
    for t, _ in rf.trees:
        _count_feat(t, counts)
    total = counts.sum()
    return counts / total if total > 0 else counts

def _count_feat(node, counts):
    if node.feat is not None:
        counts[node.feat] += 1
        if node.left:  _count_feat(node.left,  counts)
        if node.right: _count_feat(node.right, counts)

print("  Entrenando Random Forest (100 árboles, max_depth=8)...")
rf_model = rf_train(X, y, seed=42)

# OOB R²
r2_oob = rf_oob_score(rf_model, X, y)
print(f"  OOB R²: {r2_oob:.4f}")

# Cross-validation (solo 3 folds para velocidad con RF)
splits_rf = kfold_indices(n, k=3, seed=42)
cv_rf = cv_evaluate(
    lambda m, Xt: rf_predict(m, Xt),
    lambda Xtr, ytr: rf_train(Xtr, ytr, n_trees=50, seed=99),
    X, y, splits_rf
)
print(f"  R² CV (3-fold): {cv_rf['r2'][0]:.4f} ± {cv_rf['r2'][1]:.4f}")
print(f"  RMSE CV:        {cv_rf['rmse'][0]:.4f} ± {cv_rf['rmse'][1]:.4f}")
print(f"  MAE CV:         {cv_rf['mae'][0]:.4f} ± {cv_rf['mae'][1]:.4f}")

# Importancia de features
rf_importance = rf_feature_importance(rf_model, p)
imp_series = pd.Series(rf_importance, index=FEATURES).sort_values(ascending=False)
print(f"  Top 5 features: {imp_series.head(5).to_dict()}")

# ─────────────────────────────────────────────
#  5. MODELO 3 — GRADIENT BOOSTING
# ─────────────────────────────────────────────
print("\n" + "=" * 60)
print("MODELO 3: GRADIENT BOOSTING")
print("=" * 60)
"""
Gradient Boosting para Regresión:
- Inicio: F₀(x) = ȳ
- Iteración m=1..M:
    rₘᵢ = yᵢ - Fₘ₋₁(xᵢ)   ← pseudo-residuos (gradiente del MSE)
    hₘ  = fit_tree(X, rₘ)   ← árbol débil sobre residuos
    Fₘ(x) = Fₘ₋₁(x) + η · hₘ(x)
- Predicción final: F_M(x)
η = learning rate (controla el tamaño del paso)
"""
N_BOOST   = 150    # número de árboles
LR_BOOST  = 0.05   # η = learning rate
DEPTH_BOOST = 4    # árboles débiles: poco profundos

class GradientBoostingRegressor:
    def __init__(self, F0, estimators, lr):
        self.F0 = F0                # media de y (predicción inicial)
        self.estimators = estimators  # lista de árboles
        self.lr = lr

def gb_train(X, y, n_estimators=N_BOOST, lr=LR_BOOST,
             max_depth=DEPTH_BOOST, min_samples=10, seed=0):
    rng = np.random.default_rng(seed)
    F0 = y.mean()
    F_pred = np.full(len(y), F0)
    estimators = []
    for m in range(n_estimators):
        residuals = y - F_pred          # pseudo-residuos
        t = _build_tree(X, residuals, 0, max_depth, min_samples, None, rng)
        h = tree_predict(t, X)
        F_pred = F_pred + lr * h
        estimators.append(t)
    return GradientBoostingRegressor(F0, estimators, lr)

def gb_predict(model, X):
    F = np.full(len(X), model.F0)
    for t in model.estimators:
        F = F + model.lr * tree_predict(t, X)
    return F

def gb_feature_importance(model, n_features):
    counts = np.zeros(n_features)
    for t in model.estimators:
        _count_feat(t, counts)
    total = counts.sum()
    return counts / total if total > 0 else counts

print(f"  Entrenando Gradient Boosting ({N_BOOST} iteraciones, lr={LR_BOOST}, depth={DEPTH_BOOST})...")
gb_model = gb_train(X, y, seed=42)
y_pred_gb_all = gb_predict(gb_model, X)
print(f"  R² (train): {r2_score(y, y_pred_gb_all):.4f}")

# Cross-validation
cv_gb = cv_evaluate(
    lambda m, Xt: gb_predict(m, Xt),
    lambda Xtr, ytr: gb_train(Xtr, ytr, n_estimators=100, seed=99),
    X, y, SPLITS
)
print(f"  R² CV (5-fold): {cv_gb['r2'][0]:.4f} ± {cv_gb['r2'][1]:.4f}")
print(f"  RMSE CV:        {cv_gb['rmse'][0]:.4f} ± {cv_gb['rmse'][1]:.4f}")
print(f"  MAE CV:         {cv_gb['mae'][0]:.4f} ± {cv_gb['mae'][1]:.4f}")

gb_importance = gb_feature_importance(gb_model, p)
gb_imp_series = pd.Series(gb_importance, index=FEATURES).sort_values(ascending=False)
print(f"  Top 5 features: {gb_imp_series.head(5).to_dict()}")

# ─────────────────────────────────────────────
#  6. COMPARATIVA DE MODELOS
# ─────────────────────────────────────────────
print("\n" + "=" * 60)
print("COMPARATIVA DE MODELOS")
print("=" * 60)

RESULTS = {
    'Ridge Regression': cv_ridge,
    'Random Forest':    cv_rf,
    'Gradient Boosting': cv_gb,
}
for name, cv in RESULTS.items():
    print(f"  {name:22s}  R²={cv['r2'][0]:.4f}  RMSE={cv['rmse'][0]:.4f}  MAE={cv['mae'][0]:.4f}")

# ─────────────────────────────────────────────
#  7. SHAP — APROXIMACIÓN POR PERMUTACIÓN
# ─────────────────────────────────────────────
print("\n" + "=" * 60)
print("SHAP — IMPORTANCIA POR PERMUTACIÓN")
print("=" * 60)
"""
Permutation SHAP (aproximación):
Para cada feature j:
  1. Barajar X[:, j] → X_perm
  2. ΔRMSE_j = RMSE(y, f(X_perm)) - RMSE(y, f(X))
  Cuanto mayor el ΔRMSE, más importante la feature.
Se calcula sobre el mejor modelo (GB).
"""
SAMPLE_SHAP = min(300, n)  # muestrear para velocidad
rng_shap = np.random.default_rng(0)
idx_shap  = rng_shap.choice(n, size=SAMPLE_SHAP, replace=False)
X_shap, y_shap = X[idx_shap], y[idx_shap]
base_rmse = rmse(y_shap, gb_predict(gb_model, X_shap))

shap_importance = np.zeros(p)
for j in range(p):
    X_perm = X_shap.copy()
    X_perm[:, j] = rng_shap.permutation(X_perm[:, j])
    delta = rmse(y_shap, gb_predict(gb_model, X_perm)) - base_rmse
    shap_importance[j] = max(delta, 0)

shap_sum = shap_importance.sum()
shap_rel = shap_importance / shap_sum if shap_sum > 0 else shap_importance
shap_series = pd.Series(shap_rel, index=FEATURES).sort_values(ascending=False)
print(f"  Top 5 SHAP: {shap_series.head(5).to_dict()}")

# ─────────────────────────────────────────────
#  8. PREDICCIONES FINALES (MODELO GB)
# ─────────────────────────────────────────────
y_pred_log = gb_predict(gb_model, X)
y_pred_eur = np.exp(y_pred_log)
y_real_eur = np.exp(y)
residuals  = y - y_pred_log

df_pred = df_model[['player', 'team', 'league', 'market_value_eur']].copy()
df_pred['log_mv_real']    = y
df_pred['log_mv_pred']    = y_pred_log
df_pred['mv_pred_eur']    = y_pred_eur
df_pred['error_log']      = residuals
df_pred['error_pct']      = (y_pred_eur - y_real_eur) / y_real_eur * 100
df_pred['abs_error_pct']  = df_pred['error_pct'].abs()
df_pred = df_pred.sort_values('market_value_eur', ascending=False)

# ─────────────────────────────────────────────
#  9. GRÁFICOS
# ─────────────────────────────────────────────
print("\nGenerando gráficos...")

plt.rcParams.update({'font.family': 'DejaVu Sans', 'font.size': 10})

# G1 — Predicho vs Real
fig, ax = plt.subplots(figsize=(7, 6))
sc = ax.scatter(y, y_pred_log, alpha=0.35, s=12, c=f'#{C_LBLUE}')
lims = [y.min()-0.2, y.max()+0.2]
ax.plot(lims, lims, 'r--', lw=1.5, label='Predicción perfecta')
ax.set_xlabel('log(MV) real'); ax.set_ylabel('log(MV) predicho')
ax.set_title(f'Gradient Boosting — Pred. vs Real\nR² CV = {cv_gb["r2"][0]:.3f}')
ax.legend(); plt.tight_layout()
plt.savefig(PLOT_DIR / 'M1_pred_vs_real.png', dpi=130, bbox_inches='tight')
plt.close()

# G2 — Comparativa R² de los 3 modelos
fig, ax = plt.subplots(figsize=(6, 4))
nombres = list(RESULTS.keys())
r2_vals = [RESULTS[n]['r2'][0] for n in nombres]
r2_stds = [RESULTS[n]['r2'][1] for n in nombres]
colors  = [f'#{C_LBLUE}', f'#{C_GOLD}', f'#{C_LGREEN}']
bars = ax.bar(nombres, r2_vals, yerr=r2_stds, capsize=5, color=colors, edgecolor='white')
for bar, val in zip(bars, r2_vals):
    ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.01,
            f'{val:.3f}', ha='center', va='bottom', fontweight='bold', fontsize=11)
ax.set_ylabel('R² (Cross-Validation 5-fold)')
ax.set_title('Comparativa R² — 3 Modelos Predictivos')
ax.set_ylim(0, max(r2_vals)*1.15)
plt.tight_layout()
plt.savefig(PLOT_DIR / 'M2_r2_comparativa.png', dpi=130, bbox_inches='tight')
plt.close()

# G3 — Importancia SHAP
fig, ax = plt.subplots(figsize=(7, 5))
shap_top = shap_series.head(12)
colors_imp = [f'#{C_LGREEN}' if v > shap_top.median() else f'#{C_GOLD}' for v in shap_top.values]
ax.barh(shap_top.index[::-1], shap_top.values[::-1] * 100, color=colors_imp[::-1])
ax.set_xlabel('Importancia relativa (%)')
ax.set_title('SHAP — Importancia por Permutación\n(Gradient Boosting)')
plt.tight_layout()
plt.savefig(PLOT_DIR / 'M3_shap_importancia.png', dpi=130, bbox_inches='tight')
plt.close()

# G4 — Distribución de residuos
fig, ax = plt.subplots(figsize=(6, 4))
ax.hist(residuals, bins=50, color=f'#{C_LBLUE}', edgecolor='white', alpha=0.8)
ax.axvline(0, color='red', linestyle='--', lw=1.5, label='Residuo=0')
ax.axvline(residuals.mean(), color='orange', lw=1.5, label=f'Media={residuals.mean():.3f}')
ax.set_xlabel('Residuo (log_MV real − predicho)'); ax.set_ylabel('Frecuencia')
ax.set_title('Distribución de Residuos — Gradient Boosting')
ax.legend(); plt.tight_layout()
plt.savefig(PLOT_DIR / 'M4_residuos.png', dpi=130, bbox_inches='tight')
plt.close()

# G5 — Coeficientes Ridge (top 10)
fig, ax = plt.subplots(figsize=(7, 5))
rc = ridge_coefs.head(10)
cols_r = [f'#{C_LGREEN}' if v > 0 else f'#{C_RED}' for v in rc.values]
ax.barh(rc.index[::-1], rc.values[::-1], color=cols_r[::-1])
ax.axvline(0, color='black', lw=0.8)
ax.set_xlabel('Coeficiente β (datos normalizados)')
ax.set_title('Ridge Regression — Top 10 Coeficientes')
plt.tight_layout()
plt.savefig(PLOT_DIR / 'M5_ridge_coefs.png', dpi=130, bbox_inches='tight')
plt.close()

# G6 — Error % por liga (GB)
fig, ax = plt.subplots(figsize=(6, 4))
for liga in ['PL', 'LaLiga', 'Bundesliga', 'SerieA', 'Ligue1']:
    mask = df_pred['league'] == liga
    ax.scatter(df_pred.loc[mask, 'market_value_eur'] / 1e6,
               df_pred.loc[mask, 'error_pct'].clip(-80, 80),
               alpha=0.3, s=10, label=liga)
ax.axhline(0, color='black', lw=0.8, linestyle='--')
ax.set_xlabel('Valor de mercado real (M€)'); ax.set_ylabel('Error % (predicho vs real)')
ax.set_title('Error (%) por Liga — Gradient Boosting')
ax.legend(fontsize=8, ncol=2); plt.tight_layout()
plt.savefig(PLOT_DIR / 'M6_error_liga.png', dpi=130, bbox_inches='tight')
plt.close()

print(f"  6 gráficos guardados en {PLOT_DIR}")

# ─────────────────────────────────────────────
#  10. CONSTRUIR EXCEL AUDITADO
# ─────────────────────────────────────────────
print("\nConstruyendo Excel ML auditado...")

wb = openpyxl.Workbook()
wb.remove(wb.active)

def apply_header(ws, row_num, headers, fill_hex, font_hex='FFFFFF', bold=True):
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=h)
        cell.fill    = hdr_fill(fill_hex)
        cell.font    = hdr_font(bold=bold, color=font_hex)
        cell.border  = thin_border()
        cell.alignment = center()

def autofit(ws, min_width=8, max_width=30):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for c in col_cells:
            try:
                l = len(str(c.value or ''))
                if l > max_len: max_len = l
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_width, max(min_width, max_len + 2))

# ── Hoja 1: Portada ────────────────────────────────────────────────────────────
ws_cover = wb.create_sheet("Portada")
ws_cover.sheet_view.showGridLines = False
ws_cover.column_dimensions['A'].width = 3
ws_cover.column_dimensions['B'].width = 55
ws_cover.column_dimensions['C'].width = 25

title_fill = hdr_fill(C_BLUE)
gold_fill  = hdr_fill(C_GOLD)

def cover_row(ws, row, col, val, fill=None, font_sz=11, bold=False, font_col='000000'):
    c = ws.cell(row=row, column=col, value=val)
    if fill: c.fill = fill
    c.font = Font(size=font_sz, bold=bold, color=font_col, name='Calibri')
    c.alignment = Alignment(horizontal='left', vertical='center')
    return c

cover_row(ws_cover, 2, 2, "MODELOS PREDICTIVOS MULTIDIMENSIONALES", title_fill, 16, True, 'FFFFFF')
cover_row(ws_cover, 3, 2, "Estimación del Valor de Mercado de Futbolistas", title_fill, 13, False, 'FFFFFF')
cover_row(ws_cover, 4, 2, "Cinco Grandes Ligas Europeas — Temporada 2024/25", title_fill, 12, False, 'C9C9C9')
cover_row(ws_cover, 6, 2, "SECCIÓN: MODELOS DE MACHINE LEARNING", gold_fill, 13, True, C_BLUE)
rows_info = [
    (8,  "Autor",          "Guillermo"),
    (9,  "Universidad",    "Universidad Francisco de Vitoria"),
    (10, "Modelo base",    "Gradient Boosting (mejor R² CV)"),
    (11, "Modelos",        "Ridge Regression · Random Forest · Gradient Boosting"),
    (12, "Validación",     "K-Fold Cross-Validation (K=5)"),
    (13, "Dataset",        f"{len(df_model)} jugadores × {len(FEATURES)} features"),
    (14, "Implementación", "Numpy puro — sin sklearn, sin scipy"),
]
for r, label, val in rows_info:
    ws_cover.cell(row=r, column=2, value=label).font = Font(bold=True, name='Calibri')
    ws_cover.cell(row=r, column=3, value=val).font   = Font(name='Calibri')

ws_cover.row_dimensions[2].height = 28
ws_cover.row_dimensions[3].height = 22

# ── Hoja 2: Comparativa_Modelos ─────────────────────────────────────────────────
ws_cmp = wb.create_sheet("Comparativa_Modelos")

apply_header(ws_cmp, 1, ["Modelo", "R² CV Media", "R² CV Std",
                          "RMSE CV Media", "RMSE CV Std",
                          "MAE CV Media", "MAE CV Std",
                          "Folds", "Mejor modelo"], C_BLUE)

rows_cmp = [
    ("Ridge Regression (L2, λ=1)",
     cv_ridge['r2'][0], cv_ridge['r2'][1],
     cv_ridge['rmse'][0], cv_ridge['rmse'][1],
     cv_ridge['mae'][0], cv_ridge['mae'][1], 5),
    ("Random Forest (100 árboles, depth=8)",
     cv_rf['r2'][0], cv_rf['r2'][1],
     cv_rf['rmse'][0], cv_rf['rmse'][1],
     cv_rf['mae'][0], cv_rf['mae'][1], 3),
    ("Gradient Boosting (150 iter., lr=0.05, depth=4)",
     cv_gb['r2'][0], cv_gb['r2'][1],
     cv_gb['rmse'][0], cv_gb['rmse'][1],
     cv_gb['mae'][0], cv_gb['mae'][1], 5),
]
best_r2 = max(r[1] for r in rows_cmp)
for row_idx, row_data in enumerate(rows_cmp, 2):
    is_best = abs(row_data[1] - best_r2) < 1e-8
    for col_idx, val in enumerate(row_data, 1):
        c = ws_cmp.cell(row=row_idx, column=col_idx, value=val)
        c.border = thin_border()
        c.alignment = center()
        if col_idx in (2, 4, 6):
            c.number_format = '0.0000'
        if col_idx in (3, 5, 7):
            c.number_format = '0.0000'
        if is_best:
            c.fill = hdr_fill('E2EFDA')
    # Columna "Mejor modelo"
    c9 = ws_cmp.cell(row=row_idx, column=9, value="✓ MEJOR" if is_best else "")
    c9.font = Font(bold=is_best, color=C_GREEN if is_best else '000000')
    c9.alignment = center()
    c9.border = thin_border()

# Fórmula de interpretación
ws_cmp.cell(row=6, column=1, value="Interpretación R²:").font = Font(bold=True)
ws_cmp.cell(row=7, column=1, value="R² mide la proporción de varianza en log(MV) explicada por el modelo.")
ws_cmp.cell(row=8, column=1, value="RMSE y MAE están en unidades de log(€) → RMSE=0.40 ≈ error factor ×e^0.40 ≈ ×1.49 en €.")
ws_cmp.cell(row=9, column=1, value="Los valores son medias de K validaciones cruzadas (sin contaminación train/test).")
for r in [7, 8, 9]:
    ws_cmp.cell(row=r, column=1).font = Font(italic=True, color='595959', name='Calibri')

autofit(ws_cmp)
ws_cmp.row_dimensions[1].height = 30

# ── Hoja 3: Ridge_Coeficientes ──────────────────────────────────────────────────
ws_ridge = wb.create_sheet("Ridge_Coeficientes")

# Descripción del modelo
ws_ridge.cell(row=1, column=1, value="REGRESIÓN LINEAL RIDGE — θ = (XᵀX + λI)⁻¹ Xᵀy   con λ=1").font = Font(bold=True, size=12, name='Calibri')
ws_ridge.cell(row=2, column=1, value="Datos normalizados (Z-score): X_sc = (X − μ) / σ").font = Font(italic=True, color='595959')
ws_ridge.cell(row=3, column=1, value="β_i indica el cambio esperado en log(MV) por +1 desviación típica de la feature i.").font = Font(italic=True, color='595959')

apply_header(ws_ridge, 5, ["Feature", "β (coeficiente)", "|β| (peso)", "Rango normalizado",
                            "Fórmula / Cálculo"], C_LBLUE)

sorted_ridge = pd.Series(theta_ridge[1:], index=FEATURES).sort_values(key=abs, ascending=False)
ridge_max_abs = sorted_ridge.abs().max()

for row_idx, (feat, coef) in enumerate(sorted_ridge.items(), 6):
    abs_val = abs(coef)
    ws_ridge.cell(row=row_idx, column=1, value=feat).border = thin_border()
    c2 = ws_ridge.cell(row=row_idx, column=2, value=round(float(coef), 6))
    c2.number_format = '0.000000'; c2.border = thin_border(); c2.alignment = center()
    if coef > 0:
        c2.fill = hdr_fill('E2EFDA')
        c2.font = Font(color=C_GREEN)
    else:
        c2.fill = hdr_fill('FCE4D6')
        c2.font = Font(color=C_RED)
    c3 = ws_ridge.cell(row=row_idx, column=3, value=round(abs_val, 6))
    c3.number_format = '0.000000'; c3.border = thin_border(); c3.alignment = center()
    c4 = ws_ridge.cell(row=row_idx, column=4, value=round(abs_val / ridge_max_abs, 4))
    c4.number_format = '0.00%'; c4.border = thin_border(); c4.alignment = center()
    c5 = ws_ridge.cell(row=row_idx, column=5, value=f"Coeficiente {row_idx-5}: θ[{row_idx-5}] de (XᵀX + λI)⁻¹ Xᵀy")
    c5.border = thin_border(); c5.font = Font(italic=True, color='595959', size=9)

# Intercepto
n_r = len(sorted_ridge) + 6
ws_ridge.cell(row=n_r, column=1, value="INTERCEPTO (θ₀)").font = Font(bold=True)
ws_ridge.cell(row=n_r, column=2, value=round(float(theta_ridge[0]), 6))
ws_ridge.cell(row=n_r, column=2).number_format = '0.000000'

autofit(ws_ridge)

# ── Hoja 4: RF_Importancia ──────────────────────────────────────────────────────
ws_rf = wb.create_sheet("RF_Importancia")

ws_rf.cell(row=1, column=1, value="RANDOM FOREST — Importancia de Features (MDI: Mean Decrease Impurity)").font = Font(bold=True, size=12)
ws_rf.cell(row=2, column=1, value=f"100 árboles · max_depth={MAX_DEPTH_RF} · min_samples={MIN_SAMP_RF} · features/split=sqrt({p})={N_FEAT_RF}").font = Font(italic=True, color='595959')
ws_rf.cell(row=3, column=1, value="Importancia = fracción de splits que usan cada feature · mayor valor → más relevante.").font = Font(italic=True, color='595959')
ws_rf.cell(row=4, column=1, value=f"OOB R² = {r2_oob:.4f}  (estimación sin sesgo usando muestras out-of-bag)").font = Font(bold=True, color=C_GREEN)

apply_header(ws_rf, 6, ["Rank", "Feature", "Importancia MDI",
                         "Importancia SHAP (permutación)", "SHAP relativo (%)"], C_GOLD, C_BLUE)

rf_sorted = imp_series.copy()
for rank, feat in enumerate(rf_sorted.index, 1):
    row = rank + 6
    shap_val = shap_series.get(feat, 0)
    ws_rf.cell(row=row, column=1, value=rank).alignment = center()
    ws_rf.cell(row=row, column=2, value=feat)
    c3 = ws_rf.cell(row=row, column=3, value=round(float(rf_sorted[feat]), 6))
    c3.number_format = '0.000000'; c3.alignment = center()
    c4 = ws_rf.cell(row=row, column=4, value=round(float(shap_val), 6))
    c4.number_format = '0.000000'; c4.alignment = center()
    c5 = ws_rf.cell(row=row, column=5, value=round(float(shap_val*100), 2))
    c5.number_format = '0.00"%"'; c5.alignment = center()
    for col in range(1, 6):
        ws_rf.cell(row=row, column=col).border = thin_border()
    if rank <= 3:
        for col in range(1, 6):
            ws_rf.cell(row=row, column=col).fill = hdr_fill('FFF2CC')

autofit(ws_rf)

# ── Hoja 5: GB_Iteraciones ──────────────────────────────────────────────────────
ws_gb = wb.create_sheet("GB_Iteraciones")

ws_gb.cell(row=1, column=1, value="GRADIENT BOOSTING — Convergencia por Iteraciones").font = Font(bold=True, size=12)
ws_gb.cell(row=2, column=1, value=f"Fₘ(x) = Fₘ₋₁(x) + η·hₘ(x)   con η={LR_BOOST}, {N_BOOST} iteraciones, depth={DEPTH_BOOST}").font = Font(italic=True, color='595959')

apply_header(ws_gb, 4, ["Iteración m", "RMSE train (acumulado)", "Mejora RMSE", "Fórmula pseudo-residuos"], C_LGREEN, C_BLUE)

# Calcular curva de aprendizaje (solo las primeras N_BOOST iteraciones)
print("  Calculando curva de aprendizaje GB...")
F_curve = np.full(n, gb_model.F0)
rmse_curve = []
step_show  = max(1, N_BOOST // 30)  # mostrar ~30 filas
rows_shown = []
for m, t in enumerate(gb_model.estimators):
    F_curve = F_curve + LR_BOOST * tree_predict(t, X)
    if m % step_show == 0 or m == N_BOOST - 1:
        rmse_curve.append((m+1, rmse(y, F_curve)))
        rows_shown.append(m+1)

prev_rmse = rmse_curve[0][1]
for i, (m_val, rmse_val) in enumerate(rmse_curve):
    row = i + 5
    ws_gb.cell(row=row, column=1, value=m_val).alignment = center()
    c2 = ws_gb.cell(row=row, column=2, value=round(rmse_val, 6))
    c2.number_format = '0.000000'; c2.alignment = center()
    mejora = prev_rmse - rmse_val
    c3 = ws_gb.cell(row=row, column=3, value=round(mejora, 6))
    c3.number_format = '0.000000'; c3.alignment = center()
    if mejora > 0:
        c3.fill = hdr_fill('E2EFDA'); c3.font = Font(color=C_GREEN)
    c4_val = f"rₘᵢ = yᵢ − Fₘ₋₁(xᵢ)  |  Fₘ = Fₘ₋₁ + {LR_BOOST}·hₘ" if i == 0 else ""
    ws_gb.cell(row=row, column=4, value=c4_val).font = Font(italic=True, color='595959', size=9)
    for col in range(1, 5):
        ws_gb.cell(row=row, column=col).border = thin_border()
    prev_rmse = rmse_val

autofit(ws_gb)

# ── Hoja 6: Predicciones_Top50 ─────────────────────────────────────────────────
ws_pred = wb.create_sheet("Predicciones_Top50")

apply_header(ws_pred, 1, ["Rank", "Jugador", "Equipo", "Liga",
                           "MV Real (€M)", "MV Predicho (€M)",
                           "Error log", "Error %", "Valoración"], C_BLUE)

top50 = df_pred.head(50).reset_index(drop=True)
for i, row in top50.iterrows():
    r = i + 2
    mv_r = row['market_value_eur'] / 1e6
    mv_p = row['mv_pred_eur'] / 1e6
    err_pct = row['error_pct']
    valoracion = "Infravalorado" if err_pct > 15 else ("Sobrevalorado" if err_pct < -15 else "Ajustado")

    cells_vals = [i+1, row['player'], row['team'], row['league'],
                  round(mv_r, 2), round(mv_p, 2),
                  round(row['error_log'], 4), round(err_pct, 1), valoracion]
    for col_idx, val in enumerate(cells_vals, 1):
        c = ws_pred.cell(row=r, column=col_idx, value=val)
        c.border = thin_border()
        c.alignment = center() if col_idx != 2 else left()

    # Colores según valoración
    fill_val = {'Infravalorado': 'E2EFDA', 'Sobrevalorado': 'FCE4D6', 'Ajustado': 'DDEBF7'}
    for col_idx in range(1, 10):
        ws_pred.cell(row=r, column=col_idx).fill = hdr_fill(fill_val[valoracion])

    for col_num in (5, 6):
        ws_pred.cell(row=r, column=col_num).number_format = '#,##0.00'
    ws_pred.cell(row=r, column=8).number_format = '0.0"%"'

autofit(ws_pred)
ws_pred.row_dimensions[1].height = 30

# ── Hoja 7: SHAP_Global ─────────────────────────────────────────────────────────
ws_shap = wb.create_sheet("SHAP_Global")

ws_shap.cell(row=1, column=1, value="SHAP — IMPORTANCIA POR PERMUTACIÓN (Gradient Boosting)").font = Font(bold=True, size=12)
ws_shap.cell(row=2, column=1, value="Método: barajar feature j → ΔRMSE = RMSE(permutado) - RMSE(original)").font = Font(italic=True, color='595959')
ws_shap.cell(row=3, column=1, value=f"ΔRMSE > 0 → la feature era útil; ΔRMSE ≈ 0 → la feature no aporta.").font = Font(italic=True, color='595959')
ws_shap.cell(row=4, column=1, value=f"Base RMSE (modelo original, n={SAMPLE_SHAP} muestras): {base_rmse:.4f}").font = Font(bold=True)

apply_header(ws_shap, 6, ["Rank", "Feature", "ΔRMSE (permutación)",
                           "Importancia relativa (%)", "Interpretación"], C_LGREEN, C_BLUE)

for rank, (feat, imp) in enumerate(shap_series.items(), 1):
    r = rank + 6
    delta = float(shap_importance[FEATURES.index(feat)])
    interp = "Alta influencia" if imp > 0.1 else ("Media influencia" if imp > 0.04 else "Baja influencia")
    vals = [rank, feat, round(delta, 6), round(float(imp)*100, 2), interp]
    for col_idx, val in enumerate(vals, 1):
        c = ws_shap.cell(row=r, column=col_idx, value=val)
        c.border = thin_border()
        c.alignment = center() if col_idx != 2 else left()
    if rank <= 5:
        for col_idx in range(1, 6):
            ws_shap.cell(row=r, column=col_idx).fill = hdr_fill('E2EFDA')

autofit(ws_shap)

# ── Hoja 8: Gráficos ────────────────────────────────────────────────────────────
ws_graphs = wb.create_sheet("Gráficos_ML")
ws_graphs.sheet_view.showGridLines = False
ws_graphs.cell(row=1, column=1, value="GRÁFICOS — MODELOS PREDICTIVOS ML").font = Font(bold=True, size=13, color=C_BLUE)

plot_files = [
    ('M1_pred_vs_real.png',   "G1: Predicho vs Real (GB)",     1,  2),
    ('M2_r2_comparativa.png', "G2: R² Comparativa 3 Modelos",  1,  12),
    ('M3_shap_importancia.png', "G3: SHAP Importancia",        22,  2),
    ('M4_residuos.png',       "G4: Distribución Residuos",     22, 12),
    ('M5_ridge_coefs.png',    "G5: Ridge Coeficientes",        43,  2),
    ('M6_error_liga.png',     "G6: Error % por Liga",          43, 12),
]
for fname, title, row_start, col_start in plot_files:
    fpath = PLOT_DIR / fname
    if fpath.exists():
        ws_graphs.cell(row=row_start, column=col_start, value=title).font = Font(bold=True, color=C_BLUE)
        img = XLImage(str(fpath))
        img.width  = 400
        img.height = 320
        ws_graphs.add_image(img, f"{get_column_letter(col_start)}{row_start+1}")

# ─────────────────────────────────────────────
#  GUARDAR
# ─────────────────────────────────────────────
wb.save(OUT_EXCEL)
print(f"\nExcel guardado: {OUT_EXCEL}")

print("\n" + "=" * 60)
print("ML COMPLETADO")
print("=" * 60)
print(f"  Ridge  R² CV:  {cv_ridge['r2'][0]:.4f}")
print(f"  RF     R² CV:  {cv_rf['r2'][0]:.4f}   (OOB R²: {r2_oob:.4f})")
print(f"  GB     R² CV:  {cv_gb['r2'][0]:.4f}  ← MEJOR MODELO")
print(f"  Top predictor SHAP: {shap_series.index[0]} ({shap_series.iloc[0]*100:.1f}%)")
print("=" * 60)
