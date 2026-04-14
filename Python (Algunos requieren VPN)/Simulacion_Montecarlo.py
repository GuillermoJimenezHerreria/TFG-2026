"""
SIMULACIÓN MONTE CARLO — TFG
Universidad Francisco de Vitoria | Guillermo

Objetivo: estimar la incertidumbre en las predicciones del Gradient Boosting
para cada jugador, considerando la variabilidad inherente en:
  1. Incertidumbre por lesiones  (injury_days puede variar entre temporadas)
  2. Incertidumbre por rendimiento (goals, assists, shots tienen variabilidad)
  3. Incertidumbre por contrato   (contract_years decrece y puede renovar)

Método:
  - Para cada jugador, simular N_SIM escenarios perturbando sus features
    con ruido gaussiano calibrado por la desviación típica histórica de esa
    variable en el dataset (σ_feature).
  - Pasar cada escenario por el modelo GB entrenado → distribución de log(MV).
  - Convertir a euros: exp(log_MV_sim) → distribución de MV simulado.
  - Calcular: VaR5%, mediana, VaR95%, rango de incertidumbre.

Interpretación para el TFG:
  - Jugadores con mayor rango = valor de mercado más incierto (lesión-propensos,
    rendimiento variable, contrato corto).
  - Jugadores con menor rango = valor de mercado estable y predecible.
"""

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from pathlib import Path
import sys, os
import warnings
warnings.filterwarnings('ignore')

# ═══════════════════════════════════════════════════════════════
#  CONFIGURACIÓN — LOCAL vs GOOGLE COLAB
#  En Colab: sube la carpeta 'TFG Business Analytics' a Google
#  Drive y ejecuta con Runtime → Ejecutar todo.
# ═══════════════════════════════════════════════════════════════
IN_COLAB = 'google.colab' in sys.modules
if IN_COLAB:
    DB = '/content'
    import subprocess
    subprocess.run(['pip', 'install', '-q', 'openpyxl', 'lxml'],
                   capture_output=True)
else:
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    DB = os.path.join(BASE_DIR, 'Bases de Datos')

# ─────────────────────────────────────────────
#  RUTAS
# ─────────────────────────────────────────────
BASE_EXCEL = os.path.join(DB, 'Dataset_Definitivo.xlsx')
OUT_EXCEL  = os.path.join(DB, 'MonteCarlo TFG.xlsx')
PLOT_DIR   = Path('/content/mc_plots' if IN_COLAB else os.path.join(BASE_DIR, 'mc_plots'))
PLOT_DIR.mkdir(exist_ok=True)

# Parámetros de la simulación
N_SIM   = 2000   # número de escenarios por jugador
N_TOP   = 30     # jugadores a mostrar en detalle en el Excel

# ─────────────────────────────────────────────
#  PALETA
# ─────────────────────────────────────────────
C_BLUE='1F4E79'; C_LBLUE='2E75B6'; C_GOLD='C9A84C'
C_GREEN='375623'; C_LGREEN='70AD47'; C_RED='C00000'

def hdr_fill(h): return PatternFill('solid', fgColor=h)
def hdr_font(bold=True, color='FFFFFF', sz=11):
    return Font(bold=bold, color=color, size=sz, name='Calibri')
def thin_border():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)
def center(): return Alignment(horizontal='center', vertical='center', wrap_text=True)
def left():   return Alignment(horizontal='left',   vertical='center')

# ─────────────────────────────────────────────
#  1. CARGA + MODELO GB (reutilizamos el pipeline)
# ─────────────────────────────────────────────
print("=" * 60)
print("CARGANDO DATOS Y ENTRENANDO MODELO...")
print("=" * 60)

ALIASES = {
    'Nombre':'player','Player':'player','Jugador':'player','Posición':'position','Pos':'position',
    'Equipo':'team','Squad':'team',
    'Edad':'age_fbref_raw','Age':'age_fbref_raw',
    'Año de nacimiento':'birth_year','Born':'birth_year',
    'Valor de mercado':'market_value_eur','market_value_eur':'market_value_eur',
    'Ingreso total anual del club':'club_revenue','club_revenue':'club_revenue','club_revenue_eur':'club_revenue',
    'Partidos jugados esta temporada':'games_played','MP':'games_played',
    'Minutos jugados esta temporada':'minutes_played','Min':'minutes_played',
    'Goles':'goals','Gls':'goals','Asistencias':'assists','Ast':'assists',
    'Años de contrato restantes':'contract_years','contract_years_remaining':'contract_years',
    'Cantidad total de lesiones registradas':'injury_count','injury_count':'injury_count',
    'Promedio de días de baja por lesión por temporada':'injury_days','injury_days_per_season':'injury_days',
    'match_confidence':'match_confidence',
    'Entradas ganadas (FBref)':'tackles_won','tackles_won':'tackles_won',
    'Intercepciones (FBref)':'interceptions','interceptions':'interceptions',
    'Faltas cometidas (FBref)':'fouls','fouls':'fouls',
    'Faltas recibidas (FBref)':'fouled','fouled':'fouled',
    'Centros (FBref)':'crosses','crosses':'crosses',
    'Tiros totales (FBref)':'shots','shots':'shots',
    'Tiros a puerta (FBref)':'shots_on_target','shots_on_target':'shots_on_target',
    'Precisión de tiro % (FBref)':'shot_accuracy','shot_accuracy':'shot_accuracy',
}
SHEETS = {
    'PL Players':'PL','LaLiga Players':'LaLiga',
    'Bundesliga Players':'Bundesliga','Serie A Players':'SerieA','Ligue 1 Players':'Ligue1',
}
CONF_THRESHOLD = 0.85

def to_num(s):
    try:
        return float(str(s).replace(',','.').replace('€','').replace(' ','').replace('M',''))
    except:
        return np.nan

xl = pd.ExcelFile(BASE_EXCEL)
frames = []
for sheet, league in SHEETS.items():
    df = xl.parse(sheet, header=0)
    df.columns = [ALIASES.get(str(c).strip(), str(c).strip()) for c in df.columns]
    df = df.loc[:, ~df.columns.duplicated()]   # drop duplicate cols after rename
    df['league'] = league
    if 'match_confidence' not in df.columns:
        df['match_confidence'] = 1.0
    frames.append(df)

df_all = pd.concat(frames, ignore_index=True)
NUM_COLS = ['market_value_eur','club_revenue','games_played','minutes_played',
            'goals','assists','contract_years','injury_count','injury_days',
            'tackles_won','interceptions','fouls','fouled','crosses',
            'shots','shots_on_target','shot_accuracy']
for col in NUM_COLS:
    if col in df_all.columns:
        df_all[col] = df_all[col].apply(to_num)
df_all['market_value_eur'] = pd.to_numeric(df_all['market_value_eur'], errors='coerce')
df_all['match_confidence']  = pd.to_numeric(df_all['match_confidence'], errors='coerce').fillna(1.0)
df_all.loc[df_all['match_confidence'] < CONF_THRESHOLD, 'market_value_eur'] = np.nan
df_all['birth_year'] = pd.to_numeric(df_all['birth_year'], errors='coerce')
df_all['age']        = 2025 - df_all['birth_year']

POS_MAP  = {'GK':0,'DF':1,'MF':2,'FW':3,'Portero':0,'Defensa':1,'Centrocampista':2,'Delantero':3,'GK,DF':1,'DF,MF':1,'MF,FW':2,'DF,FW':1,'FW,MF':2,'MF,DF':1,'FW,DF':1,'GK,MF':2}
LIGA_MAP = {'PL':0,'LaLiga':1,'Bundesliga':2,'SerieA':3,'Ligue1':4}
df_all['pos_num']    = df_all['position'].apply(lambda x: POS_MAP.get(str(x).strip(), np.nan) if pd.notna(x) else np.nan)
df_all['league_num'] = df_all['league'].map(LIGA_MAP)
df_valid = df_all[df_all['market_value_eur'] > 0].copy()
df_valid['log_mv'] = np.log(df_valid['market_value_eur'])

FEATURES = ['age','club_revenue','games_played','minutes_played','goals','assists',
            'contract_years','injury_count','injury_days','tackles_won',
            'interceptions','fouls','fouled','crosses','shots','shots_on_target',
            'shot_accuracy','pos_num','league_num']
FEATURES = [f for f in FEATURES if f in df_valid.columns]

df_model = df_valid[FEATURES + ['log_mv','player','team','league',
                                 'market_value_eur','position']].dropna()
print(f"  Dataset: {len(df_model)} jugadores × {len(FEATURES)} features")

X = df_model[FEATURES].values.astype(float)
y = df_model['log_mv'].values.astype(float)
n, p = X.shape

# ── Gradient Boosting (mismo que modelos_ml.py) ───────────────────────────────
class DecisionNode:
    __slots__ = ['feat','thresh','left','right','value']
    def __init__(self): self.feat=self.thresh=self.left=self.right=self.value=None

def _best_split(X, y, rng):
    n, p = X.shape
    best = {'gain': -np.inf, 'feat': None, 'thresh': None}
    base_mse = ((y - y.mean())**2).mean()
    for j in range(p):
        vals = np.unique(X[:, j])
        if len(vals) <= 1: continue
        thresholds = (vals[:-1]+vals[1:])/2
        if len(thresholds) > 15: thresholds = thresholds[::max(1,len(thresholds)//15)]
        for t in thresholds:
            mask = X[:,j] <= t
            yl, yr = y[mask], y[~mask]
            if len(yl) == 0 or len(yr) == 0: continue
            nl, nr = len(yl), len(yr)
            mse = (((yl-yl.mean())**2).sum() + ((yr-yr.mean())**2).sum()) / (nl+nr)
            gain = base_mse - mse
            if gain > best['gain']: best = {'gain':gain,'feat':j,'thresh':t}
    return best

def _build(X, y, depth, max_depth, min_s, rng):
    node = DecisionNode(); node.value = y.mean()
    if depth >= max_depth or len(y) < min_s: return node
    best = _best_split(X, y, rng)
    if best['feat'] is None: return node
    node.feat = best['feat']; node.thresh = best['thresh']
    mask = X[:,node.feat] <= node.thresh
    node.left  = _build(X[mask],  y[mask],  depth+1, max_depth, min_s, rng)
    node.right = _build(X[~mask], y[~mask], depth+1, max_depth, min_s, rng)
    return node

def _pred_one(node, x):
    if node.left is None: return node.value
    return _pred_one(node.left, x) if x[node.feat] <= node.thresh else _pred_one(node.right, x)

def tree_predict(node, X):
    return np.array([_pred_one(node, x) for x in X])

class GBModel:
    def __init__(self, F0, trees, lr): self.F0=F0; self.trees=trees; self.lr=lr

def gb_train(X, y, n_est=150, lr=0.05, max_d=4, min_s=10, seed=42):
    rng = np.random.default_rng(seed)
    F = np.full(len(y), y.mean()); trees = []
    for _ in range(n_est):
        t = _build(X, y-F, 0, max_d, min_s, rng)
        F += lr * tree_predict(t, X); trees.append(t)
    return GBModel(y.mean(), trees, lr)

def gb_predict(model, X):
    F = np.full(len(X), model.F0)
    for t in model.trees: F += model.lr * tree_predict(t, X)
    return F

print("  Entrenando Gradient Boosting (modelo base para MC)...")
gb_model = gb_train(X, y, seed=42)
y_pred_base = gb_predict(gb_model, X)
r2_base = 1 - ((y - y_pred_base)**2).sum() / ((y - y.mean())**2).sum()
print(f"  R² (train): {r2_base:.4f}")

# ─────────────────────────────────────────────
#  2. CALIBRAR RUIDO PARA LA SIMULACIÓN
# ─────────────────────────────────────────────
"""
Perturbación de features:
  - Features estocásticas (rendimiento): goals, assists, shots, shots_on_target,
    tackles_won, interceptions, fouls, fouled, crosses → σ = 30% del valor medio
    de esa feature en el dataset (ruido relativo moderado)
  - Features de riesgo (lesiones): injury_days, injury_count → σ = 50% del σ
    global de esa feature (lesiones son más impredecibles)
  - Features estructurales (no se perturban): age, club_revenue, league_num,
    pos_num, contract_years (se perturba levemente ±0.5 años para simular
    incertidumbre de renovación)
  - Restricción: features >= 0 siempre (clip a 0)
"""
print("\n  Calibrando ruido para features...")

# σ por feature en el dataset real
feat_std = X.std(axis=0)  # desviación típica real de cada feature

# Factor de ruido por tipo de feature
NOISE_FACTORS = {}
PERF_FEATURES   = ['goals','assists','shots','shots_on_target','tackles_won',
                    'interceptions','fouls','fouled','crosses','shot_accuracy',
                    'games_played','minutes_played']
RISK_FEATURES   = ['injury_days','injury_count']
STRUCT_FEATURES = ['age','club_revenue','league_num','pos_num']
FLEX_FEATURES   = ['contract_years']

for i, feat in enumerate(FEATURES):
    if feat in PERF_FEATURES:
        NOISE_FACTORS[i] = 0.30  # ±30% del σ global
    elif feat in RISK_FEATURES:
        NOISE_FACTORS[i] = 0.50  # ±50% del σ global (más inciertos)
    elif feat in FLEX_FEATURES:
        NOISE_FACTORS[i] = 0.15  # ±15% (pequeña incertidumbre de contrato)
    elif feat in STRUCT_FEATURES:
        NOISE_FACTORS[i] = 0.00  # no se perturba

for i, feat in enumerate(FEATURES):
    nf = NOISE_FACTORS.get(i, 0.30)
    print(f"    {feat:25s}: σ_base={feat_std[i]:.3f}  factor={nf:.0%}  σ_ruido={feat_std[i]*nf:.3f}")

# ─────────────────────────────────────────────
#  3. SIMULACIÓN MONTE CARLO
# ─────────────────────────────────────────────
print(f"\n  Ejecutando {N_SIM} simulaciones por jugador...")

rng_mc = np.random.default_rng(42)
n_players = len(df_model)

# Matriz de resultados: (n_players × N_SIM) en log_MV
# Para eficiencia, procesar en lotes de 50 simulaciones
mc_results = np.zeros((n_players, N_SIM))

BATCH = 200  # simular 200 jugadores×N_SIM a la vez
for sim in range(N_SIM):
    # Perturbar X con ruido gaussiano calibrado
    noise = np.zeros_like(X)
    for i_feat in range(p):
        factor = NOISE_FACTORS.get(i_feat, 0.30)
        if factor > 0:
            noise[:, i_feat] = rng_mc.normal(0, feat_std[i_feat] * factor, size=n_players)
    X_perturbed = np.clip(X + noise, 0, None)  # features ≥ 0
    # Preservar features no perturbadas exactas
    for i_feat in range(p):
        if NOISE_FACTORS.get(i_feat, 0.30) == 0:
            X_perturbed[:, i_feat] = X[:, i_feat]
    mc_results[:, sim] = gb_predict(gb_model, X_perturbed)

print("  Simulación completada.")

# ─────────────────────────────────────────────
#  4. ESTADÍSTICOS DE LA DISTRIBUCIÓN SIMULADA
# ─────────────────────────────────────────────
# En escala log(MV)
mc_p5   = np.percentile(mc_results, 5,  axis=1)
mc_p25  = np.percentile(mc_results, 25, axis=1)
mc_med  = np.percentile(mc_results, 50, axis=1)
mc_p75  = np.percentile(mc_results, 75, axis=1)
mc_p95  = np.percentile(mc_results, 95, axis=1)
mc_mean = mc_results.mean(axis=1)
mc_std  = mc_results.std(axis=1)

# En escala euros
mv_p5   = np.exp(mc_p5)
mv_p95  = np.exp(mc_p95)
mv_med  = np.exp(mc_med)
mv_mean = np.exp(mc_mean)

# Predicción base (sin ruido)
mv_base = np.exp(y_pred_base)
mv_real = np.exp(y)

# Rango de incertidumbre = P95 - P5 (en M€)
uncertainty_range = (mv_p95 - mv_p5) / 1e6

# Ratio de incertidumbre = rango / mediana
uncertainty_ratio = (mv_p95 - mv_p5) / (mv_med + 1e-9)

df_mc = df_model[['player','team','league','position',
                   'market_value_eur','injury_days','injury_count',
                   'contract_years','age']].copy()
df_mc['mv_real_M']      = mv_real / 1e6
df_mc['mv_base_M']      = mv_base / 1e6
df_mc['mv_mc_p5_M']     = mv_p5   / 1e6
df_mc['mv_mc_med_M']    = mv_med  / 1e6
df_mc['mv_mc_p95_M']    = mv_p95  / 1e6
df_mc['incertidumbre_M']= uncertainty_range
df_mc['ratio_incert']   = uncertainty_ratio
df_mc['mc_std_log']     = mc_std

# Clasificar incertidumbre
q33 = np.percentile(uncertainty_ratio, 33)
q67 = np.percentile(uncertainty_ratio, 67)
df_mc['nivel_incert'] = df_mc['ratio_incert'].apply(
    lambda x: 'Alta' if x > q67 else ('Media' if x > q33 else 'Baja'))

print(f"\n  Estadísticos de incertidumbre:")
print(f"    Rango P5-P95 medio:  {uncertainty_range.mean():.1f}M€")
print(f"    Ratio medio:         {uncertainty_ratio.mean():.3f}")
print(f"    Jugadores alta incert: {(df_mc['nivel_incert']=='Alta').sum()}")

# ─────────────────────────────────────────────
#  5. GRÁFICOS
# ─────────────────────────────────────────────
print("\n  Generando gráficos...")
plt.rcParams.update({'font.family': 'DejaVu Sans', 'font.size': 10})

# G1 — Distribución del rango de incertidumbre
fig, ax = plt.subplots(figsize=(7, 4))
ax.hist(uncertainty_range, bins=40, color=f'#{C_LBLUE}', edgecolor='white', alpha=0.85)
ax.axvline(uncertainty_range.mean(), color='red', lw=1.5, linestyle='--',
           label=f'Media = {uncertainty_range.mean():.1f}M€')
ax.axvline(np.percentile(uncertainty_range, 75), color='orange', lw=1.5, linestyle=':',
           label=f'P75 = {np.percentile(uncertainty_range,75):.1f}M€')
ax.set_xlabel('Rango de incertidumbre P5–P95 (M€)'); ax.set_ylabel('Nº jugadores')
ax.set_title('Distribución del Rango de Incertidumbre Monte Carlo\n(P95 − P5 del MV simulado por jugador)')
ax.legend(); plt.tight_layout()
plt.savefig(PLOT_DIR / 'MC1_distribucion_incertidumbre.png', dpi=130, bbox_inches='tight')
plt.close()

# G2 — Top 20 jugadores más inciertos (fan chart horizontal)
top20_unc = df_mc.nlargest(20, 'incertidumbre_M').reset_index(drop=True)
fig, ax = plt.subplots(figsize=(9, 7))
y_pos = np.arange(20)
ax.barh(y_pos, top20_unc['mv_mc_p95_M'] - top20_unc['mv_mc_p5_M'],
        left=top20_unc['mv_mc_p5_M'], color=f'#{C_LBLUE}', alpha=0.5,
        height=0.6, label='Rango P5–P95')
ax.scatter(top20_unc['mv_mc_med_M'], y_pos, color=f'#{C_GOLD}', zorder=5, s=50, label='Mediana MC')
ax.scatter(top20_unc['mv_real_M'],   y_pos, color=f'#{C_RED}',  zorder=5, s=30,
           marker='D', label='MV Real')
ax.set_yticks(y_pos)
ax.set_yticklabels(top20_unc['player'], fontsize=8)
ax.set_xlabel('Valor de Mercado (M€)')
ax.set_title('Top 20 Jugadores con Mayor Incertidumbre\n(Rango P5–P95 simulado, N=2000)')
ax.legend(fontsize=9); plt.tight_layout()
plt.savefig(PLOT_DIR / 'MC2_top20_inciertos.png', dpi=130, bbox_inches='tight')
plt.close()

# G3 — Distribución MC de un jugador concreto (el más valioso del dataset)
top_player_idx = int(df_mc['mv_real_M'].idxmax())
top_name = df_mc.loc[top_player_idx, 'player']
top_sims  = np.exp(mc_results[top_player_idx]) / 1e6
fig, ax = plt.subplots(figsize=(7, 4))
ax.hist(top_sims, bins=60, color=f'#{C_LGREEN}', edgecolor='white', alpha=0.85)
ax.axvline(df_mc.loc[top_player_idx,'mv_real_M'],  color='red', lw=2, label=f'MV Real={df_mc.loc[top_player_idx,"mv_real_M"]:.0f}M€')
ax.axvline(df_mc.loc[top_player_idx,'mv_mc_p5_M'], color='orange', lw=1.5, linestyle='--', label=f'P5={df_mc.loc[top_player_idx,"mv_mc_p5_M"]:.0f}M€')
ax.axvline(df_mc.loc[top_player_idx,'mv_mc_p95_M'],color='orange', lw=1.5, linestyle='--', label=f'P95={df_mc.loc[top_player_idx,"mv_mc_p95_M"]:.0f}M€')
ax.set_xlabel('MV Simulado (M€)'); ax.set_ylabel('Frecuencia')
ax.set_title(f'Distribución Monte Carlo — {top_name}\n({N_SIM} simulaciones)')
ax.legend(); plt.tight_layout()
plt.savefig(PLOT_DIR / 'MC3_ejemplo_jugador.png', dpi=130, bbox_inches='tight')
plt.close()

# G4 — MV real vs mediana MC (correlación)
fig, ax = plt.subplots(figsize=(6, 5))
ax.scatter(df_mc['mv_real_M'], df_mc['mv_mc_med_M'],
           alpha=0.35, s=12, color=f'#{C_LBLUE}')
lim = max(df_mc['mv_real_M'].max(), df_mc['mv_mc_med_M'].max()) * 1.05
ax.plot([0, lim], [0, lim], 'r--', lw=1.5, label='Predicción perfecta')
ax.set_xlabel('MV Real (M€)'); ax.set_ylabel('MV Mediana MC (M€)')
ax.set_title('MV Real vs. Mediana Monte Carlo')
ax.legend(); plt.tight_layout()
plt.savefig(PLOT_DIR / 'MC4_real_vs_mc.png', dpi=130, bbox_inches='tight')
plt.close()

# G5 — Incertidumbre vs MV real (scatter)
fig, ax = plt.subplots(figsize=(7, 5))
scatter_colors = {'Alta': f'#{C_RED}', 'Media': f'#{C_GOLD}', 'Baja': f'#{C_LGREEN}'}
for nivel, col in scatter_colors.items():
    mask = df_mc['nivel_incert'] == nivel
    ax.scatter(df_mc.loc[mask, 'mv_real_M'],
               df_mc.loc[mask, 'incertidumbre_M'],
               alpha=0.4, s=15, color=col, label=f'Incert. {nivel} (n={mask.sum()})')
ax.set_xlabel('MV Real (M€)'); ax.set_ylabel('Rango P5–P95 (M€)')
ax.set_title('Incertidumbre vs. Valor de Mercado Real')
ax.legend(); plt.tight_layout()
plt.savefig(PLOT_DIR / 'MC5_incert_vs_mv.png', dpi=130, bbox_inches='tight')
plt.close()

print(f"  5 gráficos guardados en {PLOT_DIR}")

# ─────────────────────────────────────────────
#  6. CONSTRUIR EXCEL AUDITADO
# ─────────────────────────────────────────────
print("\nConstruyendo Excel Monte Carlo...")

wb = openpyxl.Workbook()
wb.remove(wb.active)

def apply_header(ws, row_num, headers, fill_hex, font_col='FFFFFF'):
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=row_num, column=col_idx, value=h)
        c.fill=hdr_fill(fill_hex); c.font=hdr_font(color=font_col)
        c.border=thin_border(); c.alignment=center()

def autofit(ws, min_w=8, max_w=28):
    for col in ws.columns:
        ml = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_w, max(min_w, ml+2))

# ── Portada ──────────────────────────────────────────────────────────────────
ws_p = wb.create_sheet("Portada")
ws_p.sheet_view.showGridLines = False
ws_p.column_dimensions['B'].width = 60
for row, val, sz, bold, col in [
    (2, "SIMULACIÓN MONTE CARLO — INCERTIDUMBRE EN PREDICCIONES", 15, True, C_BLUE),
    (3, "Estimación del Valor de Mercado de Futbolistas", 12, False, C_BLUE),
    (4, "Universidad Francisco de Vitoria — Temporada 2024/25", 11, False, 'A0A0A0'),
    (6, f"N = {N_SIM} simulaciones/jugador  ·  Modelo: Gradient Boosting  ·  n = {n_players} jugadores", 11, True, C_GOLD),
]:
    c = ws_p.cell(row=row, column=2, value=val)
    c.font = Font(size=sz, bold=bold, color=col if row != 6 else C_BLUE, name='Calibri')
    if row in (2,3,4): c.fill = hdr_fill(C_BLUE)
    if row == 6: c.fill = hdr_fill(C_GOLD)

# ── Hoja 1: Metodología ──────────────────────────────────────────────────────
ws_met = wb.create_sheet("Metodología_MC")
ws_met.column_dimensions['A'].width = 30; ws_met.column_dimensions['B'].width = 65
apply_header(ws_met, 1, ["Concepto", "Descripción"], C_BLUE)
items = [
    ("Método",           f"Monte Carlo: {N_SIM} perturbaciones aleatorias de las features por jugador"),
    ("Modelo subyacente","Gradient Boosting entrenado sobre 1049 jugadores (mismas condiciones que modelos_ml.py)"),
    ("Ruido features",   "xᵢ_sim = clip(xᵢ + ε,  0)  donde ε ~ N(0, factor·σ_feature)"),
    ("Factor rendimiento","goals, assists, shots, tackles, ... → factor = 30% del σ global de la feature"),
    ("Factor lesión",    "injury_days, injury_count → factor = 50% del σ (más imprevisibles)"),
    ("Factor contrato",  "contract_years → factor = 15% (pequeña incertidumbre de renovación)"),
    ("No perturbadas",   "age, club_revenue, pos_num, league_num (estructurales, no cambian en el corto plazo)"),
    ("Salida",           "Distribución de MV simulado: exp(GB(x_sim)) → P5, P25, mediana, P75, P95"),
    ("Rango de incertidumbre", "P95 − P5 en euros (M€) — indica el margen de variación plausible del MV"),
    ("Ratio incertidumbre", "(P95 − P5) / mediana — normalizado, comparable entre jugadores de distinto valor"),
    ("Clasificación",    "Alta / Media / Baja según terciles del ratio de incertidumbre"),
    ("Uso en el TFG",    "Jugadores con alta incertidumbre: perfiles volátiles (lesivos, rendimiento variable, contrato corto)"),
]
for row_idx, (k, v) in enumerate(items, 2):
    ws_met.cell(row=row_idx, column=1, value=k).font = Font(bold=True, name='Calibri')
    ws_met.cell(row=row_idx, column=2, value=v)
    for col in (1, 2): ws_met.cell(row=row_idx, column=col).border = thin_border()

# ── Hoja 2: Parámetros_Ruido ─────────────────────────────────────────────────
ws_noise = wb.create_sheet("Parámetros_Ruido")
apply_header(ws_noise, 1, ["Feature", "σ global", "Factor ruido",
                            "σ ruido aplicado", "Tipo", "Justificación"], C_LBLUE)
tipo_map = {**{f:'Rendimiento' for f in ['goals','assists','shots','shots_on_target',
                                           'tackles_won','interceptions','fouls','fouled',
                                           'crosses','shot_accuracy','games_played','minutes_played']},
            **{f:'Lesión/Riesgo' for f in ['injury_days','injury_count']},
            **{f:'Contrato' for f in ['contract_years']},
            **{f:'Estructural (fija)' for f in ['age','club_revenue','league_num','pos_num']}}
just_map = {
    'Rendimiento':     'Varía cada temporada según forma, lesiones, rol táctico',
    'Lesión/Riesgo':   'Muy impredecible; historial no garantiza próxima temporada',
    'Contrato':        'Puede renovar o no; incertidumbre moderada',
    'Estructural (fija)': 'No cambia en el corto plazo / estructural',
}
for row_idx, feat in enumerate(FEATURES, 2):
    i_feat = FEATURES.index(feat)
    factor = NOISE_FACTORS.get(i_feat, 0.30)
    sigma_ruido = feat_std[i_feat] * factor
    tipo = tipo_map.get(feat, 'Rendimiento')
    vals = [feat, round(feat_std[i_feat], 4), f'{factor:.0%}',
            round(sigma_ruido, 4), tipo, just_map.get(tipo,'')]
    for col_idx, val in enumerate(vals, 1):
        c = ws_noise.cell(row=row_idx, column=col_idx, value=val)
        c.border = thin_border(); c.alignment = center() if col_idx > 1 else left()
    if factor == 0:
        for col_idx in range(1, 7):
            ws_noise.cell(row=row_idx, column=col_idx).fill = hdr_fill('F2F2F2')
autofit(ws_noise)

# ── Hoja 3: Resultados_Top ────────────────────────────────────────────────────
ws_res = wb.create_sheet("Resultados_Top50")
hdrs = ["Jugador","Equipo","Liga","Posición",
        "MV Real (M€)","MV Pred. Base (M€)",
        "P5 MC (M€)","Mediana MC (M€)","P95 MC (M€)",
        "Rango P5-P95 (M€)","Ratio Incert.","Nivel Incertidumbre",
        "Días lesión","Años contrato"]
apply_header(ws_res, 1, hdrs, C_BLUE)

# Top 50 por MV real
top50 = df_mc.nlargest(50, 'mv_real_M').reset_index(drop=True)
fill_incert = {'Alta': 'FCE4D6', 'Media': 'FFF2CC', 'Baja': 'E2EFDA'}
for row_idx, row in top50.iterrows():
    vals = [row['player'], row['team'], row['league'], row['position'],
            round(row['mv_real_M'],1), round(row['mv_base_M'],1),
            round(row['mv_mc_p5_M'],1), round(row['mv_mc_med_M'],1), round(row['mv_mc_p95_M'],1),
            round(row['incertidumbre_M'],1), round(row['ratio_incert'],3),
            row['nivel_incert'],
            round(row.get('injury_days',0),0), round(row.get('contract_years',0),1)]
    nivel = row['nivel_incert']
    for col_idx, val in enumerate(vals, 1):
        c = ws_res.cell(row=row_idx+2, column=col_idx, value=val)
        c.border = thin_border()
        c.alignment = center() if col_idx > 4 else left()
        c.fill = hdr_fill(fill_incert.get(nivel, 'FFFFFF'))
    for col_idx in range(5, 12):
        ws_res.cell(row=row_idx+2, column=col_idx).number_format = '#,##0.0'
autofit(ws_res)

# ── Hoja 4: Todos_Jugadores ───────────────────────────────────────────────────
ws_all = wb.create_sheet("Todos_Jugadores")
apply_header(ws_all, 1, ["Jugador","Liga","MV Real (M€)",
                          "P5 MC (M€)","Mediana MC (M€)","P95 MC (M€)",
                          "Rango (M€)","Nivel Incertidumbre"], C_LGREEN, C_BLUE)
df_sorted = df_mc.sort_values('incertidumbre_M', ascending=False).reset_index(drop=True)
for row_idx, row in df_sorted.iterrows():
    nivel = row['nivel_incert']
    vals = [row['player'], row['league'],
            round(row['mv_real_M'],1), round(row['mv_mc_p5_M'],1),
            round(row['mv_mc_med_M'],1), round(row['mv_mc_p95_M'],1),
            round(row['incertidumbre_M'],1), nivel]
    for col_idx, val in enumerate(vals, 1):
        c = ws_all.cell(row=row_idx+2, column=col_idx, value=val)
        c.border = thin_border()
        c.alignment = center() if col_idx > 2 else left()
        if col_idx == 8:
            c.fill = hdr_fill(fill_incert.get(nivel, 'FFFFFF'))
autofit(ws_all)

# ── Hoja 5: Estadísticos_Globales ─────────────────────────────────────────────
ws_stat = wb.create_sheet("Estadísticos_Globales")
apply_header(ws_stat, 1, ["Métrica", "Valor", "Unidad", "Descripción"], C_BLUE)
stats_rows = [
    ("N simulaciones/jugador",  N_SIM,      "#",   "Iteraciones de Monte Carlo por jugador"),
    ("N jugadores analizados",  n_players,  "#",   "Jugadores con todas las features disponibles"),
    ("Rango P5-P95 medio",      round(uncertainty_range.mean(),1), "M€", "Media del intervalo de incertidumbre"),
    ("Rango P5-P95 mediana",    round(np.median(uncertainty_range),1), "M€", "Mediana del intervalo de incertidumbre"),
    ("Rango P5-P95 máximo",     round(uncertainty_range.max(),1), "M€",  f"Jugador más incierto: {df_mc.loc[df_mc['incertidumbre_M'].idxmax(),'player']}"),
    ("Rango P5-P95 mínimo",     round(uncertainty_range.min(),2), "M€",  "Jugador más predecible"),
    ("Ratio incert. medio",     round(uncertainty_ratio.mean(),3), "ratio", "(P95-P5)/mediana"),
    ("Jugadores incert. Alta",  int((df_mc['nivel_incert']=='Alta').sum()),  "#", f">{q67:.3f} ratio"),
    ("Jugadores incert. Media", int((df_mc['nivel_incert']=='Media').sum()), "#", f"[{q33:.3f}, {q67:.3f}]"),
    ("Jugadores incert. Baja",  int((df_mc['nivel_incert']=='Baja').sum()),  "#", f"<{q33:.3f} ratio"),
    ("σ log(MV) medio",         round(mc_std.mean(),4),  "log€", "Desv. típica de las simulaciones en log_MV"),
]
for row_idx, (k, v, u, d) in enumerate(stats_rows, 2):
    ws_stat.cell(row=row_idx, column=1, value=k).font = Font(bold=True, name='Calibri')
    ws_stat.cell(row=row_idx, column=2, value=v)
    ws_stat.cell(row=row_idx, column=3, value=u)
    ws_stat.cell(row=row_idx, column=4, value=d)
    for col in range(1, 5): ws_stat.cell(row=row_idx, column=col).border = thin_border()
autofit(ws_stat)

# ── Hoja 6: Gráficos ─────────────────────────────────────────────────────────
ws_gr = wb.create_sheet("Gráficos_MC")
ws_gr.sheet_view.showGridLines = False
ws_gr.cell(1, 1, "GRÁFICOS — SIMULACIÓN MONTE CARLO").font = Font(bold=True, size=13, color=C_BLUE)
plots = [
    ('MC1_distribucion_incertidumbre.png', "G1: Distribución Rango Incertidumbre", 1,  2),
    ('MC2_top20_inciertos.png',            "G2: Top 20 Jugadores más Inciertos",    1,  12),
    ('MC3_ejemplo_jugador.png',            "G3: Distribución MC un Jugador",        22, 2),
    ('MC4_real_vs_mc.png',                 "G4: MV Real vs. Mediana MC",            22, 12),
    ('MC5_incert_vs_mv.png',               "G5: Incertidumbre vs. MV",             43,  2),
]
for fname, title, row_s, col_s in plots:
    fpath = PLOT_DIR / fname
    if fpath.exists():
        ws_gr.cell(row_s, col_s, title).font = Font(bold=True, color=C_BLUE)
        img = XLImage(str(fpath))
        img.width = 400; img.height = 320
        ws_gr.add_image(img, f"{get_column_letter(col_s)}{row_s+1}")

# ── GUARDAR ───────────────────────────────────────────────────────────────────
wb.save(OUT_EXCEL)
print(f"\nExcel guardado: {OUT_EXCEL}")

print("\n" + "=" * 60)
print("MONTE CARLO COMPLETADO")
print("=" * 60)
print(f"  N simulaciones:           {N_SIM} por jugador")
print(f"  Rango incert. medio:      {uncertainty_range.mean():.1f}M€")
print(f"  Jugador más incierto:     {df_mc.loc[df_mc['incertidumbre_M'].idxmax(),'player']} ({uncertainty_range.max():.1f}M€)")
print(f"  Jugador más predecible:   {df_mc.loc[df_mc['incertidumbre_M'].idxmin(),'player']} ({uncertainty_range.min():.2f}M€)")
print("=" * 60)
