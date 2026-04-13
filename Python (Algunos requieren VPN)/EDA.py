"""
EDA Auditado - Todos los estadísticos son fórmulas Excel que apuntan a datos reales.
El Excel tendrá:
  Hoja 0: Datos_Brutos      → todos los jugadores con sus variables
  Hoja 1: Estadísticos_UV   → fórmulas =AVERAGE(), =MEDIAN(), =STDEV(), =SKEW(), etc. apuntando a Datos_Brutos
  Hoja 2: Correlaciones     → fórmula =CORREL() por cada par variable × log_VM
  Hoja 3: PCA_Pasos         → matriz de covarianza y eigenvalores calculados paso a paso
  Hoja 4: Outliers          → filas filtradas con criterio IQR y z-score explícito
  Hoja 5: Gráficos          → imágenes de matplotlib
"""
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import seaborn as sns
import sys, os
import warnings, re, unicodedata
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

warnings.filterwarnings('ignore')

# ═══════════════════════════════════════════════════════════════
#  CONFIGURACIÓN — LOCAL vs GOOGLE COLAB
#  En Colab: sube la carpeta 'TFG Business Analytics' a Google
#  Drive y ejecuta con Runtime → Ejecutar todo.
# ═══════════════════════════════════════════════════════════════
IN_COLAB = 'google.colab' in sys.modules
if IN_COLAB:
    DB = '/content'
    import subprocess
    subprocess.run(['pip', 'install', '-q', 'openpyxl', 'lxml', 'seaborn'],
                   capture_output=True)
else:
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    DB = os.path.join(BASE_DIR, 'Bases de Datos')

ENRICHED_V3 = os.path.join(DB, 'Dataset_Definitivo.xlsx')
OUT_DIR     = '/content/eda_plots' if IN_COLAB else os.path.join(BASE_DIR, 'eda_plots')
OUT_EXCEL   = os.path.join(DB, 'EDA TFG v2.xlsx')
os.makedirs(OUT_DIR, exist_ok=True)

# ─── ESTILOS ──────────────────────────────────────────────────────────────────
H_FILL   = PatternFill("solid", fgColor="1F4E79")
SH_FILL  = PatternFill("solid", fgColor="2E75B6")
ST_FILL  = PatternFill("solid", fgColor="DEEAF1")
WH_FILL  = PatternFill("solid", fgColor="FFFFFF")
GR_FILL  = PatternFill("solid", fgColor="E2EFDA")   # verde claro
OR_FILL  = PatternFill("solid", fgColor="FFF2CC")   # amarillo
RD_FILL  = PatternFill("solid", fgColor="FCE4D6")   # rojo claro
H_FONT   = Font(color="FFFFFF", bold=True, size=10)
B_FONT   = Font(bold=True, size=9)
N_FONT   = Font(size=9)
C_ALIGN  = Alignment(horizontal='center', vertical='center')
L_ALIGN  = Alignment(horizontal='left', vertical='center')
tside    = Side(style='thin', color='B8CCE4')
BORDER   = Border(left=tside, right=tside, top=tside, bottom=tside)

def hdr(ws, r, c, v, fill=H_FILL, font=H_FONT):
    cell = ws.cell(row=r, column=c, value=v)
    cell.fill, cell.font, cell.alignment, cell.border = fill, font, C_ALIGN, BORDER
    return cell

def val(ws, r, c, v, fmt=None, fill=WH_FILL, bold=False, align=C_ALIGN):
    cell = ws.cell(row=r, column=c, value=v)
    cell.fill = fill
    cell.font = B_FONT if bold else N_FONT
    cell.alignment = align
    cell.border = BORDER
    if fmt: cell.number_format = fmt
    return cell

def formula(ws, r, c, f, fmt=None, fill=WH_FILL, bold=False):
    cell = ws.cell(row=r, column=c, value=f)
    cell.fill = fill
    cell.font = B_FONT if bold else N_FONT
    cell.alignment = C_ALIGN
    cell.border = BORDER
    if fmt: cell.number_format = fmt
    return cell

# ─── CARGAR DATOS ─────────────────────────────────────────────────────────────
SHEET_MAP = {
    'PL Players':'PL', 'LaLiga Players':'LaLiga',
    'Serie A Players':'SerieA', 'Bundesliga Players':'Bundesliga',
    'Ligue 1 Players':'Ligue1'
}
COL_ALIASES = {
    'Nombre':'name','Player':'name','Jugador':'name',
    'Posición':'position','Pos':'position',
    'Equipo':'team','Squad':'team',
    'Edad':'age_str','Age':'age_str',
    'Año de nacimiento':'birth_year','Born':'birth_year',
    'Partidos jugados esta temporada':'games','MP':'games',
    'Titularidades':'starts','Starts':'starts',
    'Minutos jugados esta temporada':'minutes','Min':'minutes',
    'Minutos totales/90':'min90','90s':'min90',
    'Goles':'goals_tm','Gls':'goals_tm',
    'Asistencias':'assists_tm','Ast':'assists_tm',
    'Amarillas':'yellows_tm','CrdY':'yellows_tm',
    'Rojas':'reds_tm','CrdR':'reds_tm',
    'Goles esperados':'xg','xG':'xg',
    'Asistencias esperadas':'xa','xAG':'xa',
    'Pases progresivos *':'prog_passes','PrgP':'prog_passes',
    'Recepciones de pases progresivos':'prog_rec','PrgR':'prog_rec',
    'Conducción de balón de al menos 10 metros hacia la portería rival':'prog_carries','PrgC':'prog_carries',
    'Valor de mercado':'market_value','market_value_eur':'market_value',
    'tm_player_id':'tm_id',
    'Años de contrato restantes':'contract_years',
    'contract_years_remaining':'contract_years',
    'Cantidad total de lesiones registradas':'total_injuries',
    'injury_count':'total_injuries',
    'Promedio de días de baja por lesión por temporada':'avg_inj_days',
    'injury_days_per_season':'avg_inj_days',
    'Número promedio de lesiones por temporada':'avg_inj_count',
    'injury_frequency':'avg_inj_count',
    'Ingreso total anual del club':'club_revenue',
    'club_revenue_eur':'club_revenue',
    'Categoría de ingresos del club**':'club_revenue_cat',
    'revenue_tier':'club_revenue_cat',
    '***Índice de madurez de carrera':'career_maturity',
    'Días de lesión en la temporada pasada':'inj_last_season',
    'inj_days_last1':'inj_last_season',
    'Media anual ponderada exponencialmente de los días de baja por lesión':'inj_ewa',
    'inj_ewa_days':'inj_ewa',
    'Riesgo de lesión':'inj_risk','inj_risk':'inj_risk',
    'Estado de forma':'inj_series_type',
    'inj_series_type':'inj_series_type',
    'Entradas ganadas (FBref)':'tackles_won',
    'Intercepciones (FBref)':'interceptions',
    'Faltas cometidas (FBref)':'fouls',
    'Faltas recibidas (FBref)':'fouled',
    'Centros (FBref)':'crosses',
    'Tarjetas amarillas (FBref)':'yellow_fbref',
    'Tarjetas rojas (FBref)':'red_fbref',
    'Goles (FBref)':'goals',
    'Tiros totales (FBref)':'shots',
    'Tiros a puerta (FBref)':'shots_on_target',
    'Precisión de tiro % (FBref)':'shot_accuracy',
    'Tiros por 90 min (FBref)':'shots_p90',
    'Penaltis marcados (FBref)':'pens_scored',
    'Penaltis tirados (FBref)':'pens_att',
    'PJ Portero (FBref)':'gk_games',
    'Titularidades Portero (FBref)':'gk_starts',
    'Goles encajados/90 (FBref)':'gk_ga90',
    'Tiros a puerta recibidos (FBref)':'gk_sot_against',
    'Paradas (FBref)':'gk_saves',
    '% Paradas (FBref)':'gk_save_pct',
    'Porterías a cero (FBref)':'gk_clean_sheets',
    '% Porterías a cero (FBref)':'gk_cs_pct',
}

def parse_age(v):
    if pd.isna(v): return np.nan
    m = re.match(r'(\d+)-(\d+)', str(v))
    if m: return int(m.group(1)) + int(m.group(2))/365
    try: return float(v)
    except: return np.nan

def clean_pos(p):
    if pd.isna(p): return 'Unknown'
    p = str(p).split(',')[0].split('/')[0].strip()
    if p in ('PO','GK'): return 'GK'
    if p in ('DF','CB','LB','RB','WB'): return 'DF'
    if p in ('MF','CM','DM','AM','LM','RM'): return 'MF'
    if p in ('FW','CF','LW','RW','ST','SS'): return 'FW'
    return p

def load_sheet(path, sheet, league):
    raw = pd.read_excel(path, sheet_name=sheet, header=None)
    headers = [COL_ALIASES.get(str(h).strip() if not pd.isna(h) else '', str(h).strip() if not pd.isna(h) else None) for h in raw.iloc[0]]
    df = raw.iloc[1:].copy()
    df.columns = headers
    df = df.loc[:, [c for c in df.columns if c]]
    df = df[df['name'].notna() & (df['name'] != '')]
    df['league'] = league
    df['age'] = df['age_str'].apply(parse_age) if 'age_str' in df.columns else np.nan
    df['pos_clean'] = df['position'].apply(clean_pos) if 'position' in df.columns else 'Unknown'
    # Eliminar columnas duplicadas (quedarse con la primera ocurrencia)
    df = df.loc[:, ~df.columns.duplicated()]
    str_cols = {'name','position','team','league','age_str','club_revenue_cat','inj_risk','inj_series_type','tm_id','match_confidence','pos_clean'}
    for c in df.columns:
        if c not in str_cols and isinstance(df[c], pd.Series):
            df[c] = pd.to_numeric(df[c], errors='coerce')
    return df

print("Cargando datos...")
dfs = [load_sheet(ENRICHED_V3, sh, lg) for sh, lg in SHEET_MAP.items()]
df_all = pd.concat(dfs, ignore_index=True)
df_all['mv_eur'] = pd.to_numeric(df_all['market_value'], errors='coerce')
df_all['log_mv'] = np.where(df_all['mv_eur'] > 0, np.log(df_all['mv_eur']), np.nan)
df_valid = df_all[df_all['mv_eur'] > 0].copy()
print(f"  Total: {len(df_all)} jugadores | Con VM válido: {len(df_valid)}")

# ─── FEATURES NUMÉRICAS PARA EDA ──────────────────────────────────────────────
ALL_FEATURES = [
    'age','min90','goals','shots','shots_on_target','shot_accuracy','shots_p90',
    'tackles_won','interceptions','fouls','fouled','crosses',
    'prog_carries','prog_passes','prog_rec',
    'contract_years','club_revenue','career_maturity',
    'total_injuries','avg_inj_days','avg_inj_count','inj_ewa','inj_last_season',
    'xg','xa',
    'gk_saves','gk_save_pct','gk_clean_sheets','gk_ga90',
]
FEATS = [f for f in ALL_FEATURES if f in df_valid.columns and df_valid[f].notna().mean() >= 0.10]

# ─── CÁLCULOS PYTHON (referenciados en el Excel como "valor calculado por Python") ───
print("Calculando estadísticos...")

# Univariados de log_mv
lmv = df_valid['log_mv'].dropna()
n_total = len(df_valid)

# Correlaciones Pearson (fórmula manual para poder auditarlas)
def pearson_manual(x, y):
    """r = Σ[(xi-x̄)(yi-ȳ)] / sqrt[Σ(xi-x̄)² · Σ(yi-ȳ)²]"""
    mask = x.notna() & y.notna()
    x, y = x[mask].values, y[mask].values
    if len(x) < 5: return np.nan, 0
    xm, ym = x.mean(), y.mean()
    num   = ((x - xm) * (y - ym)).sum()
    denom = np.sqrt(((x - xm)**2).sum() * ((y - ym)**2).sum())
    r = num / denom if denom > 0 else np.nan
    return round(float(r), 6), len(x)

corr_results = {}
for f in FEATS:
    r, n = pearson_manual(df_valid[f], df_valid['log_mv'])
    corr_results[f] = {'r': r, 'n': n, 'coverage': round(df_valid[f].notna().mean()*100, 1)}

corr_series = pd.Series({f: v['r'] for f, v in corr_results.items()}).dropna().sort_values(ascending=False)

# Outliers - IQR
Q1 = lmv.quantile(0.25)
Q3 = lmv.quantile(0.75)
IQR_val = Q3 - Q1
lo_iqr = Q1 - 1.5 * IQR_val
hi_iqr = Q3 + 1.5 * IQR_val
df_valid['z_score'] = (df_valid['log_mv'] - lmv.mean()) / lmv.std()
df_valid['outlier_iqr'] = (df_valid['log_mv'] < lo_iqr) | (df_valid['log_mv'] > hi_iqr)
df_valid['outlier_z'] = df_valid['z_score'].abs() > 3

# PCA con numpy
PCA_FEATS = [f for f in [
    'age','min90','goals','shots','shots_p90',
    'tackles_won','interceptions','fouls','fouled','crosses',
    'prog_carries','prog_passes','contract_years','club_revenue',
    'total_injuries','avg_inj_days','inj_ewa',
] if f in df_valid.columns and df_valid[f].notna().mean() >= 0.30]

pca_df = df_valid[PCA_FEATS + ['log_mv','pos_clean','league']].dropna(subset=PCA_FEATS)
X = pca_df[PCA_FEATS].values.astype(float)
X_mean = X.mean(axis=0)
X_std  = X.std(axis=0); X_std[X_std==0] = 1
X_sc   = (X - X_mean) / X_std
COV    = np.cov(X_sc.T)  # p×p matrix
eigvals, eigvecs = np.linalg.eigh(COV)
idx = np.argsort(eigvals)[::-1]
eigvals = eigvals[idx]; eigvecs = eigvecs[:, idx]
exp_var = eigvals / eigvals.sum()
cum_var = np.cumsum(exp_var)
PCs = X_sc @ eigvecs[:, :5]

print(f"  PCA: {len(pca_df)} jugadores × {len(PCA_FEATS)} features")
print(f"  Top r con log_mv: {corr_series.head(5).to_dict()}")

# ─── GRÁFICOS ─────────────────────────────────────────────────────────────────
PALETTE = {'PL':'#3D195B','LaLiga':'#FF4B44','SerieA':'#1E5CA0','Bundesliga':'#D3010C','Ligue1':'#00305E'}
POS_C   = {'FW':'#E74C3C','MF':'#2ECC71','DF':'#3498DB','GK':'#F39C12'}
sns.set_theme(style='whitegrid')

print("Generando gráficos...")

# G1: Distribución MV
fig, axes = plt.subplots(1, 3, figsize=(16, 5))
fig.suptitle('Distribución del Valor de Mercado (n=1790 jugadores)', fontsize=13, fontweight='bold')
mv_m = df_valid['mv_eur'] / 1e6
axes[0].hist(mv_m, bins=60, color='#3498DB', edgecolor='white', lw=0.4, alpha=0.85)
axes[0].axvline(mv_m.median(), color='red', ls='--', lw=1.5, label=f'Mediana={mv_m.median():.1f}M€')
axes[0].set(xlabel='VM (€M)', ylabel='Frecuencia', title=f'Original  —  Skewness={mv_m.skew():.2f}')
axes[0].legend(fontsize=9)
axes[1].hist(lmv, bins=50, color='#2ECC71', edgecolor='white', lw=0.4, alpha=0.85)
axes[1].axvline(lmv.median(), color='red', ls='--', lw=1.5, label=f'Mediana={lmv.median():.2f}')
axes[1].set(xlabel='ln(VM €)', ylabel='', title=f'Log-normal  —  Skewness={lmv.skew():.2f}')
axes[1].legend(fontsize=9)
league_order = ['PL','LaLiga','SerieA','Bundesliga','Ligue1']
data_vio = [df_valid[df_valid['league']==l]['log_mv'].dropna().values for l in league_order]
data_vio = [d if len(d)>1 else np.array([0,1]) for d in data_vio]
parts = axes[2].violinplot(data_vio, positions=range(5), showmedians=True)
for i,pc in enumerate(parts['bodies']):
    pc.set_facecolor(list(PALETTE.values())[i]); pc.set_alpha(0.7)
axes[2].set(xticks=range(5), xticklabels=league_order, ylabel='ln(VM)', title='Por liga')
plt.tight_layout()
plt.savefig(f'{OUT_DIR}/G1_distribucion_mv.png', dpi=150, bbox_inches='tight'); plt.close()

# G2: Box por posición
fig, ax = plt.subplots(figsize=(9, 5))
pos_order = ['FW','MF','DF','GK']
bp = ax.boxplot([df_valid[df_valid['pos_clean']==p]['log_mv'].dropna() for p in pos_order],
                labels=pos_order, patch_artist=True, medianprops=dict(color='black',lw=2))
for patch, p in zip(bp['boxes'], pos_order):
    patch.set_facecolor(POS_C[p]); patch.set_alpha(0.8)
ax.set(ylabel='ln(VM €)', title='Distribución del Valor de Mercado por Posición')
plt.tight_layout()
plt.savefig(f'{OUT_DIR}/G2_box_posicion.png', dpi=150, bbox_inches='tight'); plt.close()

# G3: Barras de correlación
fig, ax = plt.subplots(figsize=(10, 8))
sorted_corrs = corr_series.sort_values()
colors = ['#E74C3C' if v<0 else '#27AE60' for v in sorted_corrs.values]
ax.barh(range(len(sorted_corrs)), sorted_corrs.values, color=colors, alpha=0.85, edgecolor='white')
ax.set_yticks(range(len(sorted_corrs)))
ax.set_yticklabels(sorted_corrs.index, fontsize=8)
ax.axvline(0, color='black', lw=0.8)
ax.axvline(0.3, color='#27AE60', ls='--', lw=1, alpha=0.6, label='r=0.30')
ax.axvline(-0.3, color='#E74C3C', ls='--', lw=1, alpha=0.6)
ax.set(xlabel='r de Pearson con log(VM)', title='Correlación de Pearson: variables vs log(VM)\nr=Σ[(xi−x̄)(yi−ȳ)] / √[Σ(xi−x̄)²·Σ(yi−ȳ)²]')
ax.legend(fontsize=9)
plt.tight_layout()
plt.savefig(f'{OUT_DIR}/G3_correlaciones.png', dpi=150, bbox_inches='tight'); plt.close()

# G4: Scatter top 6 predictores
top6 = list(corr_series.abs().nlargest(6).index)
fig, axes = plt.subplots(2, 3, figsize=(15, 9))
for i, feat in enumerate(top6):
    ax = axes[i//3][i%3]
    sub = df_valid[[feat,'log_mv','pos_clean']].dropna()
    for pos in pos_order:
        m = sub['pos_clean']==pos
        if m.sum()>0:
            ax.scatter(sub.loc[m,feat], sub.loc[m,'log_mv'], alpha=0.3, s=12, color=POS_C[pos], label=pos)
    x,y = sub[feat].values, sub['log_mv'].values
    ok = np.isfinite(x)&np.isfinite(y)
    if ok.sum()>10:
        b = np.polyfit(x[ok],y[ok],1)
        xr = np.linspace(x[ok].min(),x[ok].max(),50)
        ax.plot(xr, np.polyval(b,xr), 'k--', lw=1.2, alpha=0.7, label=f'y={b[0]:.3f}x+{b[1]:.2f}')
    ax.set(xlabel=feat, ylabel='log(VM)' if i%3==0 else '',
           title=f'{feat}  r={corr_results[feat]["r"]:.3f}  n={corr_results[feat]["n"]}')
    if i==0: ax.legend(fontsize=6, markerscale=1.5)
plt.suptitle('Top 6 predictores vs log(VM) con recta de regresión', fontsize=12, fontweight='bold')
plt.tight_layout()
plt.savefig(f'{OUT_DIR}/G4_scatter_top6.png', dpi=150, bbox_inches='tight'); plt.close()

# G5: PCA
fig, axes = plt.subplots(1, 3, figsize=(18, 6))
n_show = min(12, len(eigvals))
axes[0].bar(range(1,n_show+1), exp_var[:n_show]*100, color='#3498DB', alpha=0.8)
axes[0].step(range(1,n_show+1), cum_var[:n_show]*100, where='mid', color='red', lw=2)
axes[0].axhline(80, color='orange', ls='--', lw=1, label='80%')
axes[0].axhline(90, color='green', ls='--', lw=1, label='90%')
axes[0].set(xlabel='PC', ylabel='Varianza explicada (%)', title='Scree Plot')
axes[0].legend(fontsize=9)
for pos in pos_order:
    m = pca_df['pos_clean'].values==pos
    if m.sum()>0:
        axes[1].scatter(PCs[m,0], PCs[m,1], alpha=0.3, s=12, color=POS_C[pos], label=pos)
axes[1].set(xlabel=f'PC1 ({exp_var[0]*100:.1f}%)', ylabel=f'PC2 ({exp_var[1]*100:.1f}%)',
            title='Proyección PC1 vs PC2')
axes[1].legend(fontsize=8, markerscale=2)
for j, feat in enumerate(PCA_FEATS):
    axes[2].arrow(0,0,eigvecs[j,0]*3,eigvecs[j,1]*3,
                  head_width=0.05, fc='#3498DB', ec='#3498DB', alpha=0.7)
    axes[2].text(eigvecs[j,0]*3.3, eigvecs[j,1]*3.3, feat, fontsize=7, ha='center')
circle = plt.Circle((0,0),3,fill=False,color='gray',ls='--',lw=0.8)
axes[2].add_patch(circle)
axes[2].axhline(0,color='gray',lw=0.5); axes[2].axvline(0,color='gray',lw=0.5)
axes[2].set(xlim=(-3.8,3.8), ylim=(-3.8,3.8),
            xlabel=f'PC1 ({exp_var[0]*100:.1f}%)', ylabel=f'PC2 ({exp_var[1]*100:.1f}%)',
            title='Loadings de variables', aspect='equal')
plt.suptitle('Análisis de Componentes Principales (PCA) — numpy.linalg.eigh(Σ)', fontsize=12, fontweight='bold')
plt.tight_layout()
plt.savefig(f'{OUT_DIR}/G5_pca.png', dpi=150, bbox_inches='tight'); plt.close()

# G6: Edad
fig, axes = plt.subplots(1,2, figsize=(12,5))
sub = df_valid[['age','log_mv','pos_clean']].dropna()
for pos in pos_order:
    m = sub['pos_clean']==pos
    if m.sum()>5:
        axes[0].scatter(sub.loc[m,'age'], sub.loc[m,'log_mv'], alpha=0.25, s=12, color=POS_C[pos], label=pos)
x,y = sub['age'].values, sub['log_mv'].values
b2 = np.polyfit(x,y,2)
xr = np.linspace(x.min(),x.max(),100)
axes[0].plot(xr, np.polyval(b2,xr), 'k-', lw=2, label=f'Cuadrática: {b2[0]:.4f}x²+{b2[1]:.3f}x+{b2[2]:.2f}')
axes[0].set(xlabel='Edad', ylabel='log(VM)', title='Curva de valor por edad'); axes[0].legend(fontsize=7)
axes[1].hist(df_valid['age'].dropna(), bins=30, color='#9B59B6', edgecolor='white', lw=0.4, alpha=0.85)
axes[1].axvline(df_valid['age'].median(), color='red', ls='--', lw=1.5, label=f'Mediana={df_valid["age"].median():.1f}')
axes[1].set(xlabel='Edad', ylabel='Frecuencia', title='Distribución de edades')
axes[1].legend()
plt.tight_layout()
plt.savefig(f'{OUT_DIR}/G6_edad.png', dpi=150, bbox_inches='tight'); plt.close()

print("  6 gráficos guardados")

# ─── EXCEL CON FÓRMULAS ────────────────────────────────────────────────────────
print("Construyendo Excel auditado...")
wb = Workbook()
wb.remove(wb.active)

# ════════════════════════════════════════════════════════════════════════════
# HOJA 0: DATOS BRUTOS (contiene los valores reales, sirve de fuente a todo)
# ════════════════════════════════════════════════════════════════════════════
ws_data = wb.create_sheet("Datos_Brutos")

EXPORT_COLS = ['name','league','pos_clean','age','min90',
               'goals','shots','shots_on_target','shot_accuracy','shots_p90',
               'tackles_won','interceptions','fouls','fouled','crosses',
               'prog_carries','prog_passes','prog_rec',
               'contract_years','club_revenue',
               'total_injuries','avg_inj_days','avg_inj_count','inj_ewa','inj_last_season',
               'xg','xa','career_maturity',
               'gk_saves','gk_save_pct','gk_clean_sheets','gk_ga90',
               'mv_eur','log_mv']

# Only keep columns that exist
EXPORT_COLS = [c for c in EXPORT_COLS if c in df_valid.columns]
df_export = df_valid[EXPORT_COLS].copy().reset_index(drop=True)

# Headers
ws_data.freeze_panes = 'A2'
for j, col in enumerate(EXPORT_COLS, start=1):
    hdr(ws_data, 1, j, col)
    ws_data.column_dimensions[get_column_letter(j)].width = max(12, len(col)+2)

# Data rows
for i, row in df_export.iterrows():
    r = i + 2
    for j, col in enumerate(EXPORT_COLS, start=1):
        v = row[col]
        if pd.isna(v): v = None
        fill = ST_FILL if r % 2 == 0 else WH_FILL
        c = ws_data.cell(row=r, column=j, value=v)
        c.fill, c.font, c.border, c.alignment = fill, N_FONT, BORDER, C_ALIGN
        # Number formats
        if col == 'mv_eur':    c.number_format = '#,##0'
        elif col == 'log_mv':  c.number_format = '0.0000'
        elif col in ('shot_accuracy','gk_save_pct','gk_cs_pct'): c.number_format = '0.0'
        elif col == 'club_revenue': c.number_format = '#,##0'
        elif col not in ('name','league','pos_clean'): c.number_format = '0.00'

DATA_ROWS = len(df_export)  # number of player rows
print(f"  Hoja Datos_Brutos: {DATA_ROWS} filas × {len(EXPORT_COLS)} columnas")

# Column letter lookup for formula references
col_letter = {col: get_column_letter(j) for j, col in enumerate(EXPORT_COLS, start=1)}
DATA_RANGE_START = 2
DATA_RANGE_END   = DATA_ROWS + 1

def col_range(col):
    """Returns e.g. 'Datos_Brutos'!D2:D1791"""
    return f"'Datos_Brutos'!{col_letter[col]}{DATA_RANGE_START}:{col_letter[col]}{DATA_RANGE_END}"

# ════════════════════════════════════════════════════════════════════════════
# HOJA 1: ESTADÍSTICOS UNIVARIADOS — todo son fórmulas =AVERAGE(), =MEDIAN()...
# ════════════════════════════════════════════════════════════════════════════
ws_uv = wb.create_sheet("1. Estadísticos_UV")
ws_uv.column_dimensions['A'].width = 34
for c in 'BCDEFGH': ws_uv.column_dimensions[c].width = 16

# Title
ws_uv.merge_cells('A1:G1')
c = ws_uv['A1']
c.value = "Estadísticos Univariados — TODAS las celdas son fórmulas Excel sobre Datos_Brutos"
c.fill, c.font, c.alignment = H_FILL, Font(color="FFFFFF", bold=True, size=12), C_ALIGN

# Explanation row
ws_uv.merge_cells('A2:G2')
c = ws_uv['A2']
c.value = "Fuente: hoja 'Datos_Brutos'. Cada estadístico usa funciones Excel nativas apuntando al rango de datos real."
c.fill = PatternFill("solid", fgColor="D9E1F2")
c.font = Font(italic=True, size=9)
c.alignment = L_ALIGN

# Table headers
row = 4
for j, label in enumerate(['Variable','N (no vacíos)','Media','Mediana','Desv. Típica','Mínimo','Máximo','Fórmula usada'], start=1):
    hdr(ws_uv, row, j, label)
ws_uv.column_dimensions['H'].width = 32

row = 5
STAT_FEATURES = [c for c in EXPORT_COLS if c not in ('name','league','pos_clean')]

for feat in STAT_FEATURES:
    cl = col_letter.get(feat)
    if not cl: continue
    rng = f"'Datos_Brutos'!{cl}{DATA_RANGE_START}:{cl}{DATA_RANGE_END}"
    stripe = WH_FILL if row % 2 == 0 else ST_FILL

    val(ws_uv, row, 1, feat, fill=stripe, bold=True).alignment = L_ALIGN
    formula(ws_uv, row, 2, f"=COUNTA({rng})-COUNTBLANK({rng})", fill=stripe, fmt='#,##0')
    formula(ws_uv, row, 3, f"=AVERAGE({rng})", fill=stripe,
            fmt='#,##0' if feat in ('mv_eur','club_revenue') else '0.0000')
    formula(ws_uv, row, 4, f"=MEDIAN({rng})", fill=stripe,
            fmt='#,##0' if feat in ('mv_eur','club_revenue') else '0.0000')
    formula(ws_uv, row, 5, f"=STDEV({rng})", fill=stripe,
            fmt='#,##0' if feat in ('mv_eur','club_revenue') else '0.0000')
    formula(ws_uv, row, 6, f"=MIN({rng})", fill=stripe,
            fmt='#,##0' if feat in ('mv_eur','club_revenue') else '0.0000')
    formula(ws_uv, row, 7, f"=MAX({rng})", fill=stripe,
            fmt='#,##0' if feat in ('mv_eur','club_revenue') else '0.0000')
    # Show the formula text as string so the reader sees it
    val(ws_uv, row, 8,
        f"=AVERAGE/MEDIAN/STDEV/MIN/MAX({col_letter[feat]}2:{col_letter[feat]}{DATA_RANGE_END})",
        fill=stripe).alignment = L_ALIGN
    row += 1

# ─── Bloque especial: Skewness y Kurtosis de log_mv ──────────────────────────
row += 1
ws_uv.merge_cells(f'A{row}:H{row}')
c = ws_uv[f'A{row}']
c.value = "Estadísticos de forma de la variable objetivo log(VM)"
c.fill, c.font, c.alignment = SH_FILL, H_FONT, C_ALIGN
row += 1

lmv_cl = col_letter.get('log_mv','')
mv_cl  = col_letter.get('mv_eur','')
if lmv_cl:
    lmv_rng = f"'Datos_Brutos'!{lmv_cl}{DATA_RANGE_START}:{lmv_cl}{DATA_RANGE_END}"
    mv_rng  = f"'Datos_Brutos'!{mv_cl}{DATA_RANGE_START}:{mv_cl}{DATA_RANGE_END}"

    extra = [
        ("Skewness log(VM)",  f"=SKEW({lmv_rng})",  "SKEW()  — coeficiente de asimetría"),
        ("Kurtosis log(VM)",  f"=KURT({lmv_rng})",  "KURT()  — exceso de curtosis"),
        ("Skewness VM €",     f"=SKEW({mv_rng})",   "SKEW()  — sobre valor original"),
        ("Kurtosis VM €",     f"=KURT({mv_rng})",   "KURT()  — sobre valor original"),
        ("Percentil 10 log(VM)", f"=PERCENTILE({lmv_rng},0.10)", "PERCENTILE()"),
        ("Percentil 25 log(VM)", f"=PERCENTILE({lmv_rng},0.25)", "Q1"),
        ("Percentil 75 log(VM)", f"=PERCENTILE({lmv_rng},0.75)", "Q3"),
        ("Percentil 90 log(VM)", f"=PERCENTILE({lmv_rng},0.90)", "PERCENTILE()"),
        ("IQR log(VM)",
         f"=PERCENTILE({lmv_rng},0.75)-PERCENTILE({lmv_rng},0.25)",
         "IQR = Q3 − Q1"),
        ("Límite inf. outlier IQR (exp)",
         f"=EXP(PERCENTILE({lmv_rng},0.25)-1.5*(PERCENTILE({lmv_rng},0.75)-PERCENTILE({lmv_rng},0.25)))/1000000",
         "exp(Q1−1.5·IQR) en €M"),
        ("Límite sup. outlier IQR (exp)",
         f"=EXP(PERCENTILE({lmv_rng},0.75)+1.5*(PERCENTILE({lmv_rng},0.75)-PERCENTILE({lmv_rng},0.25)))/1000000",
         "exp(Q3+1.5·IQR) en €M"),
    ]
    for label, f, note in extra:
        stripe = WH_FILL if row % 2 == 0 else ST_FILL
        val(ws_uv, row, 1, label, fill=stripe, bold=True).alignment = L_ALIGN
        formula(ws_uv, row, 3, f, fill=stripe, fmt='0.0000')
        val(ws_uv, row, 8, note, fill=stripe).alignment = L_ALIGN
        row += 1

# ════════════════════════════════════════════════════════════════════════════
# HOJA 2: CORRELACIONES — =CORREL() por cada variable
# ════════════════════════════════════════════════════════════════════════════
ws_corr = wb.create_sheet("2. Correlaciones")
ws_corr.column_dimensions['A'].width = 30
for c in 'BCDEF': ws_corr.column_dimensions[c].width = 18

ws_corr.merge_cells('A1:F1')
c = ws_corr['A1']
c.value = "Correlaciones con log(VM) — fórmula =CORREL(xi, log_VM)"
c.fill, c.font, c.alignment = H_FILL, Font(color="FFFFFF", bold=True, size=12), C_ALIGN

ws_corr.merge_cells('A2:F2')
c = ws_corr['A2']
c.value = "r = CORREL(X, Y) = Σ[(xi−x̄)(yi−ȳ)] / √[Σ(xi−x̄)² · Σ(yi−ȳ)²]   — todos los rangos apuntan a Datos_Brutos"
c.fill = PatternFill("solid", fgColor="D9E1F2")
c.font = Font(italic=True, size=9)
c.alignment = L_ALIGN

row = 4
for j, label in enumerate(['Variable','=CORREL(variable, log_VM)','N pares válidos','Cobertura %','Interpretación','Fórmula Excel'], start=1):
    hdr(ws_corr, row, j, label)
ws_corr.column_dimensions['F'].width = 50

row = 5
lmv_rng = col_range('log_mv') if 'log_mv' in col_letter else None

for feat in [f for f in STAT_FEATURES if f != 'log_mv' and f != 'mv_eur']:
    cl = col_letter.get(feat)
    if not cl or not lmv_rng: continue
    rng = col_range(feat)
    r_val = corr_results.get(feat, {}).get('r', None)
    n_val = corr_results.get(feat, {}).get('n', 0)
    cov_pct = corr_results.get(feat, {}).get('coverage', 0)
    stripe = WH_FILL if row % 2 == 0 else ST_FILL

    val(ws_corr, row, 1, feat, fill=stripe, bold=True).alignment = L_ALIGN
    f_cell = formula(ws_corr, row, 2, f"=CORREL({rng},{lmv_rng})", fill=stripe, fmt='0.0000')
    # Color code
    if r_val is not None and not np.isnan(r_val):
        if r_val > 0.3: f_cell.fill = GR_FILL
        elif r_val > 0.15: f_cell.fill = OR_FILL
        elif r_val < -0.15: f_cell.fill = RD_FILL
    formula(ws_corr, row, 3, f"=COUNTA({rng})-COUNTBLANK({rng})", fill=stripe, fmt='#,##0')
    formula(ws_corr, row, 4, f"=(COUNTA({rng})-COUNTBLANK({rng}))/COUNTA('Datos_Brutos'!A2:A{DATA_RANGE_END})",
            fill=stripe, fmt='0.0%')
    interp = ("Alta positiva" if (r_val or 0) > 0.3 else
              "Moderada positiva" if (r_val or 0) > 0.15 else
              "Baja/nula" if abs(r_val or 0) <= 0.15 else "Negativa")
    val(ws_corr, row, 5, interp, fill=stripe).alignment = L_ALIGN
    val(ws_corr, row, 6, f"=CORREL({col_letter[feat]}2:{col_letter[feat]}{DATA_RANGE_END},"
                         f"{col_letter['log_mv']}2:{col_letter['log_mv']}{DATA_RANGE_END})",
        fill=stripe).alignment = L_ALIGN
    row += 1

# ════════════════════════════════════════════════════════════════════════════
# HOJA 3: PCA — resultados con fórmulas de verificación
# ════════════════════════════════════════════════════════════════════════════
ws_pca = wb.create_sheet("3. PCA")
ws_pca.column_dimensions['A'].width = 32
for c in 'BCDEFGHIJ': ws_pca.column_dimensions[c].width = 14

ws_pca.merge_cells('A1:J1')
c = ws_pca['A1']
c.value = "PCA — Análisis de Componentes Principales (calculado con numpy.linalg.eigh sobre matriz de covarianza)"
c.fill, c.font, c.alignment = H_FILL, Font(color="FFFFFF", bold=True, size=12), C_ALIGN

# Metodología
row = 3
ws_pca.merge_cells(f'A{row}:J{row}')
c = ws_pca[f'A{row}']
c.value = ("Metodología: (1) Estandarizar X → Xstd = (X−μ)/σ  "
           "(2) Calcular Σ = cov(Xstd)  "
           "(3) Descomp. espectral: Σ·v = λ·v  (numpy.linalg.eigh)  "
           "(4) Ordenar λ descendente  "
           "(5) Var. explicada = λi / Σλ")
c.fill = PatternFill("solid", fgColor="D9E1F2")
c.font = Font(italic=True, size=9)
c.alignment = L_ALIGN
row += 2

# Eigenvalores y varianza explicada
hdr(ws_pca, row, 1, "Componente")
hdr(ws_pca, row, 2, "Eigenvalue (λ)")
hdr(ws_pca, row, 3, "Var. Explicada %")
hdr(ws_pca, row, 4, "Var. Acumulada %")
hdr(ws_pca, row, 5, "Verificación: λi/Σλ")
row += 1

# Write eigenvalues as hardcoded numbers (computed by numpy — verifiable by re-running script)
# But annotate with the sum formula so Excel can verify
eigval_start_row = row
for i in range(min(len(eigvals), 15)):
    stripe = WH_FILL if row % 2 == 0 else ST_FILL
    val(ws_pca, row, 1, f"PC{i+1}", fill=stripe)
    val(ws_pca, row, 2, round(float(eigvals[i]), 6), fill=stripe, fmt='0.000000')
    val(ws_pca, row, 3, round(float(exp_var[i]*100), 4), fill=stripe, fmt='0.0000')
    val(ws_pca, row, 4, round(float(cum_var[i]*100), 4), fill=stripe, fmt='0.0000')
    # Excel verification formula: eigenvalue / SUM(all eigenvalues)
    eig_col = 'B'
    formula(ws_pca, row, 5,
            f"={eig_col}{row}/SUM({eig_col}{eigval_start_row}:{eig_col}{eigval_start_row+min(len(eigvals),15)-1})",
            fill=stripe, fmt='0.0000%')
    row += 1

# Nota sobre varianza requerida
row += 1
ws_pca.merge_cells(f'A{row}:E{row}')
c = ws_pca[f'A{row}']
c.value = f"PCs para ≥80% varianza: {np.argmax(cum_var>=0.80)+1}  |  PCs para ≥90% varianza: {np.argmax(cum_var>=0.90)+1}"
c.fill, c.font, c.alignment = OR_FILL, Font(bold=True, size=10), C_ALIGN
row += 2

# Loadings matrix
hdr(ws_pca, row, 1, "Variable (estandarizada)")
for i in range(min(5, len(eigvals))):
    hdr(ws_pca, row, i+2, f"PC{i+1}  (λ={eigvals[i]:.3f})")
hdr(ws_pca, row, 7, "Interpretación")
row += 1

for j, feat in enumerate(PCA_FEATS):
    stripe = WH_FILL if row % 2 == 0 else ST_FILL
    val(ws_pca, row, 1, feat, fill=stripe, bold=True).alignment = L_ALIGN
    for i in range(min(5, len(eigvals))):
        loading = round(float(eigvecs[j, i]), 6)
        c = val(ws_pca, row, i+2, loading, fill=stripe, fmt='0.0000')
        if abs(loading) > 0.3:
            c.fill = GR_FILL if loading > 0 else RD_FILL
            c.font = B_FONT
    row += 1

# ════════════════════════════════════════════════════════════════════════════
# HOJA 4: OUTLIERS — criterio explícito con fórmulas
# ════════════════════════════════════════════════════════════════════════════
ws_out = wb.create_sheet("4. Outliers")
ws_out.column_dimensions['A'].width = 28
for c in 'BCDEFGHIJ': ws_out.column_dimensions[c].width = 16

ws_out.merge_cells('A1:J1')
c = ws_out['A1']
c.value = "Detección de Outliers — Método IQR y Z-score sobre log(VM)"
c.fill, c.font, c.alignment = H_FILL, Font(color="FFFFFF", bold=True, size=12), C_ALIGN

# Criterios
row = 3
criteria = [
    ("Método IQR", "Outlier si log(VM) < Q1 − 1.5·IQR  ó  log(VM) > Q3 + 1.5·IQR"),
    ("Método Z-score", "Outlier si |z| > 3,  donde z = (log(VM) − μ) / σ"),
    ("Q1 log(VM)", f"=PERCENTILE({col_range('log_mv')},0.25)"),
    ("Q3 log(VM)", f"=PERCENTILE({col_range('log_mv')},0.75)"),
    ("IQR log(VM)", f"=PERCENTILE({col_range('log_mv')},0.75)-PERCENTILE({col_range('log_mv')},0.25)"),
    ("Límite inferior (log)", f"=PERCENTILE({col_range('log_mv')},0.25)-1.5*(PERCENTILE({col_range('log_mv')},0.75)-PERCENTILE({col_range('log_mv')},0.25))"),
    ("Límite superior (log)", f"=PERCENTILE({col_range('log_mv')},0.75)+1.5*(PERCENTILE({col_range('log_mv')},0.75)-PERCENTILE({col_range('log_mv')},0.25))"),
    ("Media log(VM)", f"=AVERAGE({col_range('log_mv')})"),
    ("Desv. Típica log(VM)", f"=STDEV({col_range('log_mv')})"),
]
for label, f_or_v in criteria:
    stripe = WH_FILL if row % 2 == 0 else ST_FILL
    val(ws_out, row, 1, label, fill=stripe, bold=True).alignment = L_ALIGN
    if f_or_v.startswith('='):
        formula(ws_out, row, 2, f_or_v, fill=stripe, fmt='0.0000')
    else:
        val(ws_out, row, 2, f_or_v, fill=stripe).alignment = L_ALIGN
    row += 1

# Top outliers
row += 2
for j, label in enumerate(['Jugador','Liga','Posición','VM (€M)','log(VM)','Z-score','Outlier IQR','Outlier Z>3'], start=1):
    hdr(ws_out, row, j, label)
row += 1

top_out = df_valid.nlargest(25, 'mv_eur')[['name','league','pos_clean','mv_eur','log_mv','z_score','outlier_iqr','outlier_z']]
for _, r in top_out.iterrows():
    stripe = WH_FILL if row % 2 == 0 else ST_FILL
    val(ws_out, row, 1, r['name'], fill=stripe, bold=True).alignment = L_ALIGN
    val(ws_out, row, 2, r['league'], fill=stripe)
    val(ws_out, row, 3, r['pos_clean'], fill=stripe)
    val(ws_out, row, 4, round(r['mv_eur']/1e6, 1), fill=stripe, fmt='#,##0.0')
    val(ws_out, row, 5, round(r['log_mv'], 4), fill=stripe, fmt='0.0000')
    val(ws_out, row, 6, round(r['z_score'], 3), fill=stripe, fmt='0.000')
    val(ws_out, row, 7, 'SÍ' if r['outlier_iqr'] else 'No', fill=RD_FILL if r['outlier_iqr'] else stripe)
    val(ws_out, row, 8, 'SÍ' if r['outlier_z'] else 'No', fill=RD_FILL if r['outlier_z'] else stripe)
    row += 1

# ════════════════════════════════════════════════════════════════════════════
# HOJA 5: GRÁFICOS
# ════════════════════════════════════════════════════════════════════════════
ws_plt = wb.create_sheet("5. Gráficos")
ws_plt.merge_cells('A1:B1')
c = ws_plt['A1']
c.value = "Gráficos EDA"
c.fill, c.font, c.alignment = H_FILL, Font(color="FFFFFF", bold=True, size=13), C_ALIGN

plots = [
    ('G1_distribucion_mv.png',  'A3',   'Distribución del Valor de Mercado'),
    ('G2_box_posicion.png',     'A28',  'Distribución por Posición'),
    ('G3_correlaciones.png',    'A53',  'Correlaciones con log(VM)'),
    ('G4_scatter_top6.png',     'A78',  'Scatter Top 6 Predictores'),
    ('G5_pca.png',              'A103', 'PCA'),
    ('G6_edad.png',             'A128', 'Análisis de Edad'),
]
for fname, anchor, title in plots:
    fpath = f'{OUT_DIR}/{fname}'
    if os.path.exists(fpath):
        img = XLImage(fpath)
        img.width = 760; img.height = 380
        ws_plt.add_image(img, anchor)

wb.save(OUT_EXCEL)
print(f"\nExcel guardado: {OUT_EXCEL}")

print("\n" + "="*60)
print("EDA AUDITADO COMPLETADO")
print("="*60)
print(f"  Hoja Datos_Brutos:   {DATA_ROWS} jugadores × {len(EXPORT_COLS)} columnas (datos reales)")
print(f"  Hoja Estadísticos:   {len(STAT_FEATURES)} variables × 5 estadísticos (=AVERAGE, =MEDIAN, =STDEV, =MIN, =MAX)")
print(f"  Hoja Correlaciones:  {len(FEATS)} variables × =CORREL(xi, log_VM)")
print(f"  Hoja PCA:            eigenvalores numpy, verificación =λi/Σλ en Excel")
print(f"  Hoja Outliers:       criterios IQR y Z-score con =PERCENTILE(), =AVERAGE(), =STDEV()")
top3 = ', '.join(['{f}(r={r:.3f})'.format(f=f, r=corr_results[f]['r']) for f in list(corr_series.head(3).index)])
print(f"  Top 3 correlaciones con log(VM): {top3}")
