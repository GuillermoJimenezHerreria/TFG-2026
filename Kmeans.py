"""
K-MEANS CLUSTERING — TFG
Universidad Francisco de Vitoria | Guillermo

Implementación desde CERO con numpy (sin sklearn).

Objetivo: segmentar los 1049 jugadores del dataset ML en K grupos
(arquetipos) según sus características y analizar si los clusters
capturan diferencias reales de valor de mercado.

Algoritmo K-Means (Lloyd):
  1. Inicializar K centroides (K-Means++ para mayor estabilidad)
  2. Asignar cada punto al centroide más cercano (distancia euclídea)
  3. Recalcular centroides como media del cluster
  4. Repetir 2-3 hasta convergencia (cambio < ε) o max_iter
  5. Criterio de selección de K: Elbow method (inercia) + Silhouette score

Output: Excel auditado con perfiles de clusters, top jugadores por cluster,
        gráficos PCA 2D de los clusters
"""

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from pathlib import Path
import sys, os
import warnings
warnings.filterwarnings('ignore')

# ═══════════════════════════════════════════════════════════════
#  CONFIGURACIÓN — LOCAL vs GOOGLE COLAB
#  En Colab: sube la carpeta 'TFG Business Analytics' a Google
#  Drive y ejecuta con Runtime → Ejecutar todo.
# ═══════════════════════════════════════════════════════════════
IN_COLAB = 'google.colab' in sys.modules
if IN_COLAB:
    DB = '/content'
    import subprocess
    subprocess.run(['pip', 'install', '-q', 'openpyxl', 'lxml'],
                   capture_output=True)
else:
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    DB = os.path.join(BASE_DIR, 'Bases de Datos')

# ─────────────────────────────────────────────
#  RUTAS
# ─────────────────────────────────────────────
BASE_EXCEL = os.path.join(DB, 'Dataset_Definitivo.xlsx')
OUT_EXCEL  = os.path.join(DB, 'KMeans TFG.xlsx')
PLOT_DIR   = Path('/content/km_plots' if IN_COLAB else os.path.join(BASE_DIR, 'km_plots'))
PLOT_DIR.mkdir(exist_ok=True)

# ─────────────────────────────────────────────
#  PALETA
# ─────────────────────────────────────────────
C_BLUE   = '1F4E79'; C_LBLUE  = '2E75B6'; C_GOLD = 'C9A84C'
C_GREEN  = '375623'; C_LGREEN = '70AD47'; C_RED  = 'C00000'
C_GREY   = 'F2F2F2'; C_WHITE  = 'FFFFFF'
CLUSTER_COLORS = ['#E63946','#457B9D','#2A9D8F','#E9C46A','#F4A261','#264653']

def hdr_fill(h): return PatternFill('solid', fgColor=h)
def hdr_font(bold=True, color='FFFFFF', sz=11):
    return Font(bold=bold, color=color, size=sz, name='Calibri')
def thin_border():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)
def center(): return Alignment(horizontal='center', vertical='center', wrap_text=True)
def left():   return Alignment(horizontal='left',   vertical='center')

# ─────────────────────────────────────────────
#  1. CARGA DE DATOS (igual que modelos_ml.py)
# ─────────────────────────────────────────────
print("=" * 60)
print("CARGANDO DATOS...")
print("=" * 60)

ALIASES = {
    'Nombre':'player','Player':'player','Jugador':'player','Posición':'position','Pos':'position',
    'Equipo':'team','Squad':'team',
    'Edad':'age_fbref_raw','Age':'age_fbref_raw',
    'Año de nacimiento':'birth_year','Born':'birth_year',
    'Valor de mercado':'market_value_eur','market_value_eur':'market_value_eur',
    'Ingreso total anual del club':'club_revenue','club_revenue':'club_revenue','club_revenue_eur':'club_revenue',
    'Partidos jugados esta temporada':'games_played','MP':'games_played',
    'Minutos jugados esta temporada':'minutes_played','Min':'minutes_played',
    'Goles':'goals','Gls':'goals','Asistencias':'assists','Ast':'assists',
    'Años de contrato restantes':'contract_years','contract_years_remaining':'contract_years',
    'Cantidad total de lesiones registradas':'injury_count','injury_count':'injury_count',
    'Promedio de días de baja por lesión por temporada':'injury_days','injury_days_per_season':'injury_days',
    'match_confidence':'match_confidence',
    'Entradas ganadas (FBref)':'tackles_won','tackles_won':'tackles_won',
    'Intercepciones (FBref)':'interceptions','interceptions':'interceptions',
    'Faltas cometidas (FBref)':'fouls','fouls':'fouls',
    'Faltas recibidas (FBref)':'fouled','fouled':'fouled',
    'Centros (FBref)':'crosses','crosses':'crosses',
    'Tiros totales (FBref)':'shots','shots':'shots',
    'Tiros a puerta (FBref)':'shots_on_target','shots_on_target':'shots_on_target',
    'Precisión de tiro % (FBref)':'shot_accuracy','shot_accuracy':'shot_accuracy',
    'Riesgo de lesión':'injury_risk',
}
SHEETS = {
    'PL Players':'PL','LaLiga Players':'LaLiga',
    'Bundesliga Players':'Bundesliga','Serie A Players':'SerieA','Ligue 1 Players':'Ligue1',
}
CONF_THRESHOLD = 0.85

def to_num(s):
    try:
        return float(str(s).replace(',','.').replace('€','').replace(' ','').replace('M',''))
    except:
        return np.nan

xl = pd.ExcelFile(BASE_EXCEL)
frames = []
for sheet, league in SHEETS.items():
    df = xl.parse(sheet, header=0)
    df.columns = [ALIASES.get(str(c).strip(), str(c).strip()) for c in df.columns]
    df = df.loc[:, ~df.columns.duplicated()]   # drop duplicate columns after rename
    df['league'] = league
    if 'match_confidence' not in df.columns:
        df['match_confidence'] = 1.0
    frames.append(df)

df_all = pd.concat(frames, ignore_index=True)

NUM_COLS = ['market_value_eur','club_revenue','games_played','minutes_played',
            'goals','assists','contract_years','injury_count','injury_days',
            'tackles_won','interceptions','fouls','fouled','crosses',
            'shots','shots_on_target','shot_accuracy']
for col in NUM_COLS:
    if col in df_all.columns:
        df_all[col] = pd.to_numeric(df_all[col].apply(to_num), errors='coerce')

df_all['market_value_eur'] = pd.to_numeric(df_all['market_value_eur'], errors='coerce')
df_all['match_confidence']  = pd.to_numeric(df_all['match_confidence'], errors='coerce').fillna(1.0)
df_all.loc[df_all['match_confidence'] < CONF_THRESHOLD, 'market_value_eur'] = np.nan
df_all['birth_year'] = pd.to_numeric(df_all['birth_year'], errors='coerce')
df_all['age'] = 2025 - df_all['birth_year']

POS_MAP = {'GK':0,'DF':1,'MF':2,'FW':3,'Portero':0,'Defensa':1,'Centrocampista':2,'Delantero':3,
           'GK,DF':1,'DF,MF':1,'MF,FW':2,'DF,FW':1,'FW,MF':2,'MF,DF':1,'FW,DF':1,'GK,MF':2}
LIGA_MAP = {'PL':0,'LaLiga':1,'Bundesliga':2,'SerieA':3,'Ligue1':4}
df_all['pos_num']    = df_all['position'].apply(lambda x: POS_MAP.get(str(x).strip(), np.nan) if pd.notna(x) else np.nan)
df_all['league_num'] = df_all['league'].map(LIGA_MAP)

df_valid = df_all[df_all['market_value_eur'] > 0].copy()
df_valid['log_mv'] = np.log(df_valid['market_value_eur'])

FEATURES = ['age','club_revenue','games_played','minutes_played','goals','assists',
            'contract_years','injury_count','injury_days','tackles_won',
            'interceptions','fouls','fouled','crosses','shots','shots_on_target',
            'shot_accuracy','pos_num','league_num']
FEATURES = [f for f in FEATURES if f in df_valid.columns]

df_model = df_valid[FEATURES + ['log_mv','player','team','league',
                                 'market_value_eur','position']].dropna()
print(f"  Dataset: {len(df_model)} jugadores × {len(FEATURES)} features")

X = df_model[FEATURES].values.astype(float)
y = df_model['log_mv'].values.astype(float)
n, p = X.shape

# Normalización Z-score (los clusters son sensibles a la escala)
X_mean = X.mean(axis=0)
X_std  = X.std(axis=0)
X_std[X_std == 0] = 1
X_sc   = (X - X_mean) / X_std

# ─────────────────────────────────────────────
#  2. IMPLEMENTACIÓN K-MEANS DESDE CERO
# ─────────────────────────────────────────────
print("\n" + "=" * 60)
print("K-MEANS CLUSTERING (implementación numpy)")
print("=" * 60)

def kmeans_plus_plus_init(X, k, seed=42):
    """
    K-Means++ initialization (Arthur & Vassilvitskii, 2007):
    Elige los centroides iniciales con probabilidad proporcional
    a d²(x, centroide más cercano ya elegido) → menos sensible
    a inicializaciones malas que el K-Means estándar.
    """
    rng = np.random.default_rng(seed)
    n   = len(X)
    # Primer centroide: aleatorio
    centroids = [X[rng.integers(0, n)]]
    for _ in range(k - 1):
        # Distancia al centroide más cercano para cada punto
        dists = np.array([min(np.sum((x - c)**2) for c in centroids) for x in X])
        probs = dists / dists.sum()
        cumprobs = np.cumsum(probs)
        r = rng.random()
        idx = np.searchsorted(cumprobs, r)
        centroids.append(X[idx])
    return np.array(centroids)

def kmeans(X, k, max_iter=300, tol=1e-4, seed=42):
    """
    K-Means Lloyd + inicialización K-Means++
    Devuelve: labels, centroids, inertia, n_iter
    """
    centroids = kmeans_plus_plus_init(X, k, seed)
    labels = np.zeros(len(X), dtype=int)

    for it in range(max_iter):
        # Paso E: asignar cada punto al centroide más cercano
        # Distancia euclídea: ||x - c||² = ||x||² - 2xᵀc + ||c||²
        dists = (np.sum(X**2, axis=1, keepdims=True)
                 - 2 * X @ centroids.T
                 + np.sum(centroids**2, axis=1))
        new_labels = np.argmin(dists, axis=1)

        # Paso M: recalcular centroides
        new_centroids = np.array([
            X[new_labels == j].mean(axis=0) if (new_labels == j).sum() > 0
            else centroids[j]
            for j in range(k)
        ])

        # Convergencia
        shift = np.sum((new_centroids - centroids)**2)
        labels    = new_labels
        centroids = new_centroids
        if shift < tol:
            break

    # Inercia = suma de distancias² intra-cluster
    inertia = sum(np.sum((X[labels == j] - centroids[j])**2)
                  for j in range(k))
    return labels, centroids, inertia, it + 1

def silhouette_score(X, labels):
    """
    Silhouette score medio:
    s(i) = (b(i) - a(i)) / max(a(i), b(i))
    a(i) = distancia media intra-cluster
    b(i) = distancia media al cluster vecino más cercano
    [-1, 1] → 1 = clustering perfecto
    """
    k = len(np.unique(labels))
    n = len(X)
    scores = np.zeros(n)
    # Para eficiencia, calcular distancias por cluster
    for i in range(n):
        ci = labels[i]
        # a(i): distancia media a los demás puntos de su cluster
        mask_same = (labels == ci)
        mask_same[i] = False
        if mask_same.sum() == 0:
            scores[i] = 0
            continue
        a_i = np.sqrt(((X[mask_same] - X[i])**2).sum(axis=1)).mean()
        # b(i): distancia media al cluster vecino más cercano
        b_i = np.inf
        for j in range(k):
            if j == ci:
                continue
            mask_j = (labels == j)
            if mask_j.sum() == 0:
                continue
            d_j = np.sqrt(((X[mask_j] - X[i])**2).sum(axis=1)).mean()
            if d_j < b_i:
                b_i = d_j
        scores[i] = (b_i - a_i) / max(a_i, b_i) if max(a_i, b_i) > 0 else 0
    return scores.mean()

# ─────────────────────────────────────────────
#  3. ELBOW METHOD + SILHOUETTE → ELEGIR K
# ─────────────────────────────────────────────
print("  Calculando Elbow & Silhouette para K=2..8...")
K_RANGE   = range(2, 9)
inertias  = []
sil_scores = []

# Subsample para silhouette (es O(n²), costoso)
np.random.seed(0)
sil_idx = np.random.choice(n, size=min(300, n), replace=False)
X_sil   = X_sc[sil_idx]

for k in K_RANGE:
    labels_k, _, inertia_k, _ = kmeans(X_sc, k, seed=42)
    inertias.append(inertia_k)
    labels_sil = labels_k[sil_idx]
    # Solo calcular silhouette si todos los clusters están representados
    unique_in_sample = np.unique(labels_sil)
    if len(unique_in_sample) < k:
        sil_scores.append(np.nan)
    else:
        s = silhouette_score(X_sil, labels_sil)
        sil_scores.append(s)
    print(f"    K={k}: inercia={inertia_k:.1f}  silhouette={sil_scores[-1]:.4f}")

# Elegir K por silhouette máximo (desempate: menor inercia)
valid_sil = [(i, s) for i, s in enumerate(sil_scores) if not np.isnan(s)]
best_k_idx = max(valid_sil, key=lambda x: x[1])[0]
K_OPT = list(K_RANGE)[best_k_idx]
print(f"\n  K óptimo elegido: K={K_OPT} (silhouette={sil_scores[best_k_idx]:.4f})")

# ─────────────────────────────────────────────
#  4. CLUSTERING FINAL CON K ÓPTIMO
# ─────────────────────────────────────────────
print(f"\n  Ejecutando K-Means final con K={K_OPT}...")
# Ejecutar 10 veces con distintas semillas → elegir la de menor inercia
best_result = None
best_inertia = np.inf
for seed in range(10):
    labels, centroids, inertia, n_iter = kmeans(X_sc, K_OPT, seed=seed*7)
    if inertia < best_inertia:
        best_inertia = inertia
        best_result  = (labels, centroids, inertia, n_iter)

labels, centroids, inertia_final, n_iter_final = best_result
df_model = df_model.copy()
df_model['cluster'] = labels

print(f"  Convergió en {n_iter_final} iteraciones | inercia={inertia_final:.1f}")
print(f"  Distribución por cluster:")
for c in sorted(df_model['cluster'].unique()):
    n_c  = (labels == c).sum()
    mv_c = df_model.loc[df_model['cluster']==c, 'market_value_eur'].median() / 1e6
    print(f"    Cluster {c}: {n_c:4d} jugadores | MV mediana={mv_c:.1f}M€")

# ─────────────────────────────────────────────
#  5. CARACTERIZACIÓN DE CLUSTERS
# ─────────────────────────────────────────────
print("\n  Caracterizando clusters...")

# Estadísticos por cluster
cluster_stats = []
for c in sorted(df_model['cluster'].unique()):
    mask  = df_model['cluster'] == c
    sub   = df_model[mask]
    n_c   = mask.sum()
    mv_med = sub['market_value_eur'].median() / 1e6
    mv_mean= sub['market_value_eur'].mean()   / 1e6
    age_m  = sub['age'].mean()       if 'age'          in sub.columns else np.nan
    goals_m= sub['goals'].mean()     if 'goals'        in sub.columns else np.nan
    shots_m= sub['shots'].mean()     if 'shots'        in sub.columns else np.nan
    min_m  = sub['minutes_played'].mean() if 'minutes_played' in sub.columns else np.nan
    # Liga más frecuente
    liga_mode = sub['league'].mode()[0] if len(sub) > 0 else '?'
    # Posición más frecuente
    pos_mode  = sub['position'].mode()[0] if len(sub) > 0 else '?'
    cluster_stats.append({
        'cluster': c, 'n': n_c,
        'mv_mediana_M': round(mv_med, 1), 'mv_media_M': round(mv_mean, 1),
        'edad_media': round(age_m, 1), 'goles_media': round(goals_m, 2),
        'tiros_media': round(shots_m, 1), 'minutos_media': round(min_m, 0),
        'liga_predominante': liga_mode, 'posicion_predominante': pos_mode,
    })

cs_df = pd.DataFrame(cluster_stats).set_index('cluster')

# Nombrar clusters según perfil de MV + edad
def nombre_cluster(row):
    mv = row['mv_mediana_M']
    age = row['edad_media']
    goals = row['goles_media']
    shots = row['tiros_media']
    if mv >= 40:
        return 'Elite / Estrellas'
    elif mv >= 15 and age < 24:
        return 'Jóvenes con Proyección'
    elif mv >= 15:
        return 'Titulares Consolidados'
    elif shots > 15 or goals > 3:
        return 'Atacantes Medianos'
    elif age > 28:
        return 'Veteranos / Rotación'
    else:
        return 'Suplentes / Periferia'

cs_df['nombre'] = cs_df.apply(nombre_cluster, axis=1)
print("\n  Perfiles de cluster:")
print(cs_df[['nombre','n','mv_mediana_M','mv_media_M','edad_media','goles_media']].to_string())

# ─────────────────────────────────────────────
#  6. PCA 2D PARA VISUALIZAR CLUSTERS
# ─────────────────────────────────────────────
print("\n  PCA para visualización 2D...")
COV = np.cov(X_sc.T)
eigvals, eigvecs = np.linalg.eigh(COV)
idx = np.argsort(eigvals)[::-1]
eigvecs = eigvecs[:, idx]
PC = X_sc @ eigvecs[:, :2]
var_exp = eigvals[idx][:2] / eigvals.sum()

# ─────────────────────────────────────────────
#  7. GRÁFICOS
# ─────────────────────────────────────────────
print("\n  Generando gráficos...")
plt.rcParams.update({'font.family': 'DejaVu Sans', 'font.size': 10})

# G1 — Elbow method
fig, ax1 = plt.subplots(figsize=(7, 4))
ax2 = ax1.twinx()
ks = list(K_RANGE)
ax1.plot(ks, inertias, 'o-', color=f'#{C_LBLUE}', lw=2, label='Inercia (izq)')
ax1.set_ylabel('Inercia (suma dist² intra-cluster)', color=f'#{C_LBLUE}')
ax2.plot(ks, sil_scores, 's--', color=f'#{C_GOLD}', lw=2, label='Silhouette (der)')
ax2.set_ylabel('Silhouette score', color=f'#{C_GOLD}')
ax1.axvline(K_OPT, color='red', linestyle=':', lw=1.5, label=f'K óptimo={K_OPT}')
ax1.set_xlabel('Número de clusters K')
ax1.set_title('Elbow Method + Silhouette Score')
lines1, labels1 = ax1.get_legend_handles_labels()
lines2, labels2 = ax2.get_legend_handles_labels()
ax1.legend(lines1+lines2, labels1+labels2, fontsize=9, loc='upper right')
plt.tight_layout()
plt.savefig(PLOT_DIR / 'K1_elbow_silhouette.png', dpi=130, bbox_inches='tight')
plt.close()

# G2 — PCA 2D scatter por cluster
fig, ax = plt.subplots(figsize=(8, 6))
for c in range(K_OPT):
    mask = labels == c
    nombre_c = cs_df.loc[c, 'nombre']
    ax.scatter(PC[mask, 0], PC[mask, 1],
               alpha=0.45, s=20, color=CLUSTER_COLORS[c % len(CLUSTER_COLORS)],
               label=f'C{c}: {nombre_c} (n={mask.sum()})')
# Centroides en espacio PCA
centroids_pca = centroids @ eigvecs[:, :2]
for c in range(K_OPT):
    ax.scatter(centroids_pca[c, 0], centroids_pca[c, 1],
               color=CLUSTER_COLORS[c % len(CLUSTER_COLORS)],
               s=180, marker='*', edgecolors='black', linewidths=0.8, zorder=5)
ax.set_xlabel(f'PC1 ({var_exp[0]*100:.1f}% var. explicada)')
ax.set_ylabel(f'PC2 ({var_exp[1]*100:.1f}% var. explicada)')
ax.set_title(f'K-Means K={K_OPT} — Proyección PCA 2D\n(★ = centroide del cluster)')
ax.legend(fontsize=8, loc='best')
plt.tight_layout()
plt.savefig(PLOT_DIR / 'K2_pca_clusters.png', dpi=130, bbox_inches='tight')
plt.close()

# G3 — Boxplot MV por cluster
fig, ax = plt.subplots(figsize=(8, 5))
data_by_c = [df_model.loc[df_model['cluster']==c, 'market_value_eur'].values / 1e6
             for c in range(K_OPT)]
labels_c  = [f"C{c}\n{cs_df.loc[c,'nombre']}" for c in range(K_OPT)]
bp = ax.boxplot(data_by_c, labels=labels_c, patch_artist=True,
                medianprops={'color':'black','lw':2})
for patch, color in zip(bp['boxes'], CLUSTER_COLORS):
    patch.set_facecolor(color); patch.set_alpha(0.6)
ax.set_ylabel('Valor de mercado (M€)')
ax.set_title('Distribución de Valor de Mercado por Cluster')
ax.set_ylim(0, df_model['market_value_eur'].max() / 1e6 * 1.1)
plt.xticks(rotation=15, ha='right')
plt.tight_layout()
plt.savefig(PLOT_DIR / 'K3_mv_boxplot.png', dpi=130, bbox_inches='tight')
plt.close()

# G4 — Radar / Heatmap de centroides normalizados (top 8 features)
TOP_FEATS = ['age','club_revenue','games_played','minutes_played',
             'goals','assists','shots','contract_years']
TOP_FEATS = [f for f in TOP_FEATS if f in FEATURES]
tf_idx    = [FEATURES.index(f) for f in TOP_FEATS]
# Centroides en espacio original (des-normalizado, luego re-normalizado por feature)
cent_raw = centroids[:, tf_idx] * X_std[tf_idx] + X_mean[tf_idx]
cent_norm = (cent_raw - cent_raw.min(axis=0)) / (cent_raw.max(axis=0) - cent_raw.min(axis=0) + 1e-9)

fig, ax = plt.subplots(figsize=(9, 5))
im = ax.imshow(cent_norm, cmap='RdYlGn', aspect='auto', vmin=0, vmax=1)
ax.set_xticks(range(len(TOP_FEATS)))
ax.set_xticklabels(TOP_FEATS, rotation=35, ha='right', fontsize=9)
ax.set_yticks(range(K_OPT))
ax.set_yticklabels([f"C{c}: {cs_df.loc[c,'nombre']}" for c in range(K_OPT)], fontsize=9)
for i in range(K_OPT):
    for j in range(len(TOP_FEATS)):
        ax.text(j, i, f'{cent_raw[i, j]:.1f}', ha='center', va='center',
                fontsize=7.5, color='black')
plt.colorbar(im, ax=ax, label='Valor normalizado [0-1]')
ax.set_title('Heatmap de Centroides por Cluster (valores originales)')
plt.tight_layout()
plt.savefig(PLOT_DIR / 'K4_heatmap_centroides.png', dpi=130, bbox_inches='tight')
plt.close()

# G5 — Composición de posiciones por cluster
fig, ax = plt.subplots(figsize=(8, 5))
pos_order = ['GK','DF','MF','FW','Portero','Defensa','Centrocampista','Delantero']
for c in range(K_OPT):
    sub = df_model[df_model['cluster']==c]
    pos_counts = sub['position'].value_counts()
    # Simplificar a 4 posiciones
    simple = {'GK':0,'DF':0,'MF':0,'FW':0}
    for pos, cnt in pos_counts.items():
        p = str(pos)
        if 'GK' in p or 'Portero' in p:         simple['GK'] += cnt
        elif 'DF' in p or 'Defensa' in p:        simple['DF'] += cnt
        elif 'MF' in p or 'Centrocampista' in p: simple['MF'] += cnt
        elif 'FW' in p or 'Delantero' in p:      simple['FW'] += cnt
    total = sum(simple.values())
    if total > 0:
        x = np.arange(4) + c * 0.15
        vals = [simple[p] / total * 100 for p in ['GK','DF','MF','FW']]
ax.clear()
width = 0.15
pos_names = ['GK','DF','MF','FW']
x = np.arange(4)
for c in range(K_OPT):
    sub = df_model[df_model['cluster']==c]
    simple = {'GK':0,'DF':0,'MF':0,'FW':0}
    for pos, cnt in sub['position'].value_counts().items():
        p = str(pos)
        if 'GK' in p or 'Portero' in p:         simple['GK'] += cnt
        elif 'DF' in p or 'Defensa' in p:        simple['DF'] += cnt
        elif 'MF' in p or 'Centrocampista' in p: simple['MF'] += cnt
        elif 'FW' in p or 'Delantero' in p:      simple['FW'] += cnt
    total = sum(simple.values()) or 1
    vals = [simple[p] / total * 100 for p in pos_names]
    ax.bar(x + c * width, vals, width, label=f'C{c}: {cs_df.loc[c,"nombre"]}',
           color=CLUSTER_COLORS[c % len(CLUSTER_COLORS)], alpha=0.8)
ax.set_xticks(x + width * (K_OPT-1) / 2)
ax.set_xticklabels(pos_names)
ax.set_ylabel('% de jugadores en el cluster')
ax.set_title('Composición por Posición en cada Cluster')
ax.legend(fontsize=8, loc='upper right')
plt.tight_layout()
plt.savefig(PLOT_DIR / 'K5_posiciones_cluster.png', dpi=130, bbox_inches='tight')
plt.close()

print(f"  5 gráficos guardados en {PLOT_DIR}")

# ─────────────────────────────────────────────
#  8. CONSTRUIR EXCEL AUDITADO
# ─────────────────────────────────────────────
print("\nConstruyendo Excel K-Means auditado...")

wb = openpyxl.Workbook()
wb.remove(wb.active)

def apply_header(ws, row_num, headers, fill_hex, font_col='FFFFFF'):
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=row_num, column=col_idx, value=h)
        c.fill = hdr_fill(fill_hex); c.font = hdr_font(color=font_col)
        c.border = thin_border(); c.alignment = center()

def autofit(ws, min_w=8, max_w=28):
    for col in ws.columns:
        ml = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_w, max(min_w, ml+2))

# ── Portada ──────────────────────────────────────────────────────────────────
ws_p = wb.create_sheet("Portada")
ws_p.sheet_view.showGridLines = False
ws_p.column_dimensions['B'].width = 55
for row, val, sz, bold, col in [
    (2, "K-MEANS CLUSTERING — SEGMENTACIÓN DE JUGADORES", 15, True, C_BLUE),
    (3, "Modelos Predictivos Multidimensionales para la Estimación del Valor de Mercado", 12, False, C_BLUE),
    (4, "Universidad Francisco de Vitoria — Temporada 2024/25", 11, False, 'A0A0A0'),
    (6, f"K óptimo: {K_OPT} clusters  |  Algoritmo: K-Means++ (Lloyd)  |  n={len(df_model)} jugadores", 11, True, C_GOLD),
]:
    c = ws_p.cell(row=row, column=2, value=val)
    c.font = Font(size=sz, bold=bold, color=col, name='Calibri')
    c.fill = hdr_fill(C_BLUE) if row in (2,3,4) else hdr_fill(C_GOLD) if row == 6 else PatternFill()

# ── Hoja 1: Metodología ──────────────────────────────────────────────────────
ws_met = wb.create_sheet("Metodología")
ws_met.column_dimensions['A'].width = 25
ws_met.column_dimensions['B'].width = 70
apply_header(ws_met, 1, ["Concepto", "Descripción"], C_BLUE)
items = [
    ("Algoritmo",        "K-Means (Lloyd) con inicialización K-Means++"),
    ("Inicialización",   "K-Means++: centroide i se elige con P ∝ d²(x, centroide más cercano ya elegido)"),
    ("Paso E (asignar)", "label(xᵢ) = argmin_j ||xᵢ − cⱼ||²  [distancia euclídea al cuadrado]"),
    ("Paso M (actualizar)", "cⱼ = (1/|Cⱼ|) Σᵢ∈Cⱼ xᵢ  [media de los puntos asignados al cluster j]"),
    ("Convergencia",     "Parar si Σⱼ ||cⱼ_nuevo − cⱼ_viejo||² < ε = 1e-4  o max_iter = 300"),
    ("Inercia",          "W(K) = Σⱼ Σᵢ∈Cⱼ ||xᵢ − cⱼ||²  [suma total de distancias intra-cluster al cuadrado]"),
    ("Silhouette s(i)",  "s(i) = (b(i) − a(i)) / max(a(i), b(i))  donde a=dist. media intra, b=dist. media inter más cercana"),
    ("Selección de K",   f"Máximo silhouette sobre K=2..8  →  K óptimo = {K_OPT}"),
    ("Escalado",         "Z-score: X_sc = (X − μ) / σ  [necesario: K-Means sensible a la escala]"),
    ("Robustez",         "10 ejecuciones con semillas distintas → se elige la de menor inercia final"),
    ("Features usadas",  ", ".join(FEATURES)),
    ("Normalización MV", f"Solo jugadores con match_confidence ≥ {CONF_THRESHOLD}"),
]
for row_idx, (k, v) in enumerate(items, 2):
    ws_met.cell(row=row_idx, column=1, value=k).font = Font(bold=True, name='Calibri')
    ws_met.cell(row=row_idx, column=2, value=v).font = Font(name='Calibri')
    for col in (1, 2):
        ws_met.cell(row=row_idx, column=col).border = thin_border()

# ── Hoja 2: Elbow_Silhouette ──────────────────────────────────────────────────
ws_el = wb.create_sheet("Elbow_Silhouette")
apply_header(ws_el, 1, ["K", "Inercia W(K)",
                         "ΔInercia (W(K-1)−W(K))", "Silhouette Score",
                         "K seleccionado", "Fórmula"], C_LBLUE)
prev_in = None
for row_idx, (k, inertia, sil) in enumerate(zip(K_RANGE, inertias, sil_scores), 2):
    delta = (prev_in - inertia) if prev_in is not None else np.nan
    is_opt = (k == K_OPT)
    vals = [k, round(inertia, 1), round(delta, 1) if not np.isnan(delta) else '',
            round(sil, 4) if not np.isnan(sil) else '',
            "✓ ÓPTIMO" if is_opt else ""]
    formula_txt = "W(K) = Σⱼ Σᵢ∈Cⱼ ||xᵢ−cⱼ||²" if row_idx == 2 else ""
    vals.append(formula_txt)
    for col_idx, val in enumerate(vals, 1):
        c = ws_el.cell(row=row_idx, column=col_idx, value=val)
        c.border = thin_border(); c.alignment = center()
        if is_opt:
            c.fill = hdr_fill('E2EFDA')
            c.font = Font(bold=True, color=C_GREEN, name='Calibri')
    prev_in = inertia
autofit(ws_el)

# ── Hoja 3: Perfiles_Clusters ──────────────────────────────────────────────────
ws_cl = wb.create_sheet("Perfiles_Clusters")
hdrs = ["Cluster", "Nombre / Arquetipo", "N jugadores",
        "MV Mediana (M€)", "MV Media (M€)", "Edad Media",
        "Goles/j media", "Tiros/j media", "Minutos media",
        "Liga predominante", "Posición predominante"]
apply_header(ws_cl, 1, hdrs, C_GOLD, C_BLUE)

for row_idx, c in enumerate(sorted(df_model['cluster'].unique()), 2):
    row = cs_df.loc[c]
    vals = [c, row['nombre'], row['n'],
            row['mv_mediana_M'], row['mv_media_M'], row['edad_media'],
            row['goles_media'], row['tiros_media'], int(row['minutos_media']),
            row['liga_predominante'], row['posicion_predominante']]
    for col_idx, val in enumerate(vals, 1):
        cell = ws_cl.cell(row=row_idx, column=col_idx, value=val)
        cell.border = thin_border(); cell.alignment = center()
        cell.fill = hdr_fill(CLUSTER_COLORS[c].lstrip('#'))
        cell.font = Font(bold=(col_idx <= 2), color='000000', name='Calibri')
    for col_idx in (4, 5):
        ws_cl.cell(row=row_idx, column=col_idx).number_format = '#,##0.0'
autofit(ws_cl)

# ── Hoja 4: Top10_por_Cluster ──────────────────────────────────────────────────
ws_top = wb.create_sheet("Top10_por_Cluster")
apply_header(ws_top, 1, ["Cluster", "Arquetipo", "Jugador", "Equipo", "Liga",
                          "MV Real (M€)", "Edad", "Goles", "Tiros"], C_BLUE)
current_row = 2
for c in sorted(df_model['cluster'].unique()):
    top10 = (df_model[df_model['cluster'] == c]
             .nlargest(10, 'market_value_eur')
             .reset_index(drop=True))
    for _, player_row in top10.iterrows():
        vals = [c, cs_df.loc[c, 'nombre'],
                player_row['player'], player_row['team'], player_row['league'],
                round(player_row['market_value_eur']/1e6, 1),
                int(player_row['age']) if pd.notna(player_row.get('age')) else '',
                round(player_row.get('goals', 0), 1),
                round(player_row.get('shots', 0), 1)]
        for col_idx, val in enumerate(vals, 1):
            cell = ws_top.cell(row=current_row, column=col_idx, value=val)
            cell.border = thin_border()
            cell.alignment = center() if col_idx != 3 else left()
            cell.fill = hdr_fill(CLUSTER_COLORS[c].lstrip('#') + '40'
                                  if len(CLUSTER_COLORS[c].lstrip('#')) == 6
                                  else 'F2F2F2')
        ws_top.cell(current_row, 6).number_format = '#,##0.0'
        current_row += 1
    # Separador entre clusters
    for col_idx in range(1, 10):
        ws_top.cell(row=current_row, column=col_idx).fill = hdr_fill('D9D9D9')
    current_row += 1
autofit(ws_top)

# ── Hoja 5: Dataset_con_Clusters ──────────────────────────────────────────────
ws_dat = wb.create_sheet("Dataset_Clusters")
EXPORT_COLS = ['player','team','league','position','market_value_eur',
               'age','goals','assists','shots','contract_years','cluster']
EXPORT_COLS = [c for c in EXPORT_COLS if c in df_model.columns]
apply_header(ws_dat, 1, EXPORT_COLS, C_BLUE)
for row_idx, (_, row) in enumerate(df_model[EXPORT_COLS].iterrows(), 2):
    for col_idx, col_name in enumerate(EXPORT_COLS, 1):
        val = row[col_name]
        if col_name == 'market_value_eur':
            val = round(float(val)/1e6, 2) if pd.notna(val) else ''
        c = ws_dat.cell(row=row_idx, column=col_idx, value=val)
        c.border = thin_border()
        c.alignment = center() if col_idx > 4 else left()
    if row['cluster'] in range(len(CLUSTER_COLORS)):
        color = CLUSTER_COLORS[int(row['cluster'])].lstrip('#')
        ws_dat.cell(row=row_idx, column=EXPORT_COLS.index('cluster')+1).fill = hdr_fill(color)
ws_dat.cell(1, EXPORT_COLS.index('market_value_eur')+1).value += ' (M€)'
autofit(ws_dat)

# ── Hoja 6: Gráficos ─────────────────────────────────────────────────────────
ws_gr = wb.create_sheet("Gráficos_KMeans")
ws_gr.sheet_view.showGridLines = False
ws_gr.cell(1, 1, "GRÁFICOS — K-MEANS CLUSTERING").font = Font(bold=True, size=13, color=C_BLUE)
plots = [
    ('K1_elbow_silhouette.png', "G1: Elbow + Silhouette",  1, 2),
    ('K2_pca_clusters.png',     "G2: PCA 2D por Cluster",   1, 12),
    ('K3_mv_boxplot.png',       "G3: MV por Cluster",       22, 2),
    ('K4_heatmap_centroides.png',"G4: Heatmap Centroides",  22, 12),
    ('K5_posiciones_cluster.png',"G5: Posiciones por Cluster",43, 2),
]
for fname, title, row_s, col_s in plots:
    fpath = PLOT_DIR / fname
    if fpath.exists():
        ws_gr.cell(row_s, col_s, title).font = Font(bold=True, color=C_BLUE)
        img = XLImage(str(fpath))
        img.width = 400; img.height = 320
        ws_gr.add_image(img, f"{get_column_letter(col_s)}{row_s+1}")

# ── GUARDAR ───────────────────────────────────────────────────────────────────
wb.save(OUT_EXCEL)
print(f"\nExcel guardado: {OUT_EXCEL}")

print("\n" + "=" * 60)
print("K-MEANS COMPLETADO")
print("=" * 60)
print(f"  K óptimo: {K_OPT}  |  Inercia final: {inertia_final:.1f}")
print(f"  Silhouette score: {sil_scores[best_k_idx]:.4f}")
for c in sorted(df_model['cluster'].unique()):
    row = cs_df.loc[c]
    print(f"  Cluster {c} [{row['nombre']:28s}]: n={row['n']:4d}  MV_med={row['mv_mediana_M']:.1f}M€")
print("=" * 60)
